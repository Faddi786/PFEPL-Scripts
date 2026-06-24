"""
Fill Tender_Form_Template from cleaned_tenderlist_standardized.
- Only fills existing data rows in the template (no extra rows added).
- Only maps cleaned columns that exist in the template; extra cleaned columns are omitted.
- If cleaned has more rows than the template has data rows, those extra cleaned rows are omitted.
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path

# Paths
SCRIPT_DIR = Path(__file__).resolve().parent
INPUT_DIR = SCRIPT_DIR / "input"
OUTPUT_DIR = SCRIPT_DIR / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
CLEANED_PATH = INPUT_DIR / "cleaned_tenderlist.xlsx"
TEMPLATE_PATH = INPUT_DIR / "tender_form_template.xlsx"
OUTPUT_PATH = OUTPUT_DIR / "tender_form_filled.xlsx"
OUTPUT_PATH_ALT = OUTPUT_DIR / "tender_form_filled_alt.xlsx"

# Map: cleaned column name (normalized) -> template column index (0-based)
# Template row 1 = field labels, so data goes in same column index
CLEANED_TO_TEMPLATE = {
    "Sr. No.": 1,  # Internal Project No
    "TYPE OF TENDER": 2,
    "Name of Work": 0,  # also 44 in PROJECT DETAILS
    "Project Value \n(In Cr.)": 8,
    "Tender Documents Fee": 9,
    "Form of\nTender Documents Fee": 10,
    "EMD \n(In Cr.)": 11,
    "Form of\nEMD": 12,
    "Completion period": 26,
    "Pre Bid Meeting": 17,
    "Last Date of Submission": 22,
    "Bid Opening date": 19,
    "physical Submission": 18,  # Physical Document Submission Due Date
    "SELF/JV": 20,
    "Depertment": 3,
    "Tender ID": 7,
    "Location": 58,  # Project Remarks (or 27 Remarks)
    "BG Request issued to A/C": 14,
    "Category of Works": None,  # appended to Remarks below
    "Physical Submission": 21,
    "Remarks": 27,
}


def clean_value(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    s = str(val).strip()
    return s


def main():
    # Load cleaned list (header in row 1)
    cleaned = pd.read_excel(CLEANED_PATH, sheet_name=0, header=1)
    # Drop completely empty rows
    cleaned = cleaned.dropna(how="all")
    # Drop header-like row if present (e.g. "Sr. No." in first cell)
    if len(cleaned) and str(cleaned.iloc[0, 0]).strip() == "Sr. No.":
        cleaned = cleaned.iloc[1:].reset_index(drop=True)

    # Load template with openpyxl to keep format
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active
    # Template: row 1 = section headers, row 2 = field labels, then data rows
    first_data_row = 3
    template_data_row_count = max(0, ws.max_row - (first_data_row - 1))  # number of existing data rows

    # Use only as many cleaned rows as the template has data rows; omit extra cleaned rows
    cleaned_to_use = cleaned.iloc[:template_data_row_count]
    if len(cleaned) > template_data_row_count:
        print(f"Omitting {len(cleaned) - template_data_row_count} cleaned row(s) (template has {template_data_row_count} data row(s)).")

    # Fill only existing template rows (no extra rows added)
    for idx, row in cleaned_to_use.iterrows():
        data_row = first_data_row + idx

        for cleaned_col_name, template_col in CLEANED_TO_TEMPLATE.items():
            if template_col is None or cleaned_col_name not in cleaned.columns:
                continue
            val = row.get(cleaned_col_name)
            val = clean_value(val)
            if val:
                # openpyxl is 1-based
                ws.cell(row=data_row, column=template_col + 1, value=val)

        # Duplicate "Name of Work" into PROJECT DETAILS (col 44) if present
        if "Name of Work" in cleaned.columns:
            name_val = clean_value(row.get("Name of Work"))
            if name_val:
                ws.cell(row=data_row, column=45, value=name_val)  # 44 -> 45 1-based

        # Combine Location and Category of Works into Remarks when present
        rem_col = 28  # Remarks = template col 27 -> 1-based 28
        loc = clean_value(row.get("Location")) if "Location" in cleaned.columns else ""
        rem = clean_value(row.get("Remarks")) if "Remarks" in cleaned.columns else ""
        cat = clean_value(row.get("Category of Works")) if "Category of Works" in cleaned.columns else ""
        parts = []
        if loc:
            parts.append(f"Location: {loc}")
        if cat:
            parts.append(f"Category: {cat}")
        if rem:
            parts.append(rem)
        if parts:
            ws.cell(row=data_row, column=rem_col, value=" | ".join(parts))

    # Save (use alternate path if main is open)
    try:
        wb.save(OUTPUT_PATH)
        print(f"Filled {len(cleaned_to_use)} tender row(s) into {OUTPUT_PATH} (template has {template_data_row_count} data row(s), no extra rows added)")
    except PermissionError:
        wb.save(OUTPUT_PATH_ALT)
        print(f"Filled {len(cleaned_to_use)} tender row(s) into {OUTPUT_PATH_ALT} (template has {template_data_row_count} data row(s); output was open, saved as _2)")


if __name__ == "__main__":
    main()
