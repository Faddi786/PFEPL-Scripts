from __future__ import annotations

import math
import re
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.neural_network import MLPRegressor
from sklearn.preprocessing import StandardScaler


ROOT_DIR = Path(__file__).resolve().parents[1]
REQUIRED_DIR = ROOT_DIR / "Required Station Info"
OUTPUT_DIR = ROOT_DIR / "Filled Values"
NEIGHBOR_REPORT = ROOT_DIR / "Excel Files" / "station_neighbors_report.xlsx"
FINAL_MAPPED_FILE = ROOT_DIR / "Excel Files" / "Final_Mapped.xlsx"
ERROR_SUMMARY_FILE = OUTPUT_DIR / "Error_Summary_All_Stations.xlsx"

PRIMARY_RADIUS_KM = 30.0
FALLBACK_RADIUS_KM = 50.0
MIN_TRAIN_ROWS = 60
HOLDOUT_FRACTION = 0.20
TEMPORAL_INTERP_MAX_GAP_DAYS = 7


def _normalize_name(name: str) -> str:
    return " ".join(str(name).strip().split()).casefold()


def _remove_brackets(name: str) -> str:
    return re.sub(r"\([^)]*\)", "", str(name)).strip()


def _alnum_key(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", _normalize_name(name))


def _name_keys(name: str) -> list[str]:
    keys = [
        _normalize_name(name),
        _normalize_name(_remove_brackets(name)),
        _alnum_key(name),
        _alnum_key(_remove_brackets(name)),
    ]
    return list(dict.fromkeys([k for k in keys if k]))


def _build_mlp_model() -> MLPRegressor:
    return MLPRegressor(
        hidden_layer_sizes=(32, 16),
        activation="relu",
        solver="adam",
        alpha=1e-4,
        learning_rate_init=1e-3,
        max_iter=250,
        early_stopping=True,
        validation_fraction=0.1,
        n_iter_no_change=15,
        random_state=42,
    )


def _read_station_sheet(path: Path, sheet_name: str) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=sheet_name)
    if "Date" not in df.columns or "Rainfall Value" not in df.columns:
        return pd.DataFrame(columns=["Date", "Rainfall Value"])
    out = df[["Date", "Rainfall Value"]].copy()
    out["Date"] = pd.to_datetime(out["Date"], errors="coerce")
    out["Rainfall Value"] = pd.to_numeric(out["Rainfall Value"], errors="coerce")
    out = out.dropna(subset=["Date"]).drop_duplicates(subset=["Date"], keep="first").sort_values("Date")
    return out.reset_index(drop=True)


def _load_all_station_series(required_dir: Path) -> tuple[dict[str, pd.Series], dict[str, str], dict[str, str]]:
    """
    Returns:
      series_by_key: resolved_key -> rainfall series indexed by date
      display_by_key: resolved_key -> station display name
      workbook_by_key: resolved_key -> workbook name (group)
    """
    series_by_key: dict[str, pd.Series] = {}
    display_by_key: dict[str, str] = {}
    workbook_by_key: dict[str, str] = {}

    for wb in sorted(required_dir.glob("*.xlsx")):
        if wb.name.startswith("~$"):
            continue
        xls = pd.ExcelFile(wb)
        for sheet in xls.sheet_names:
            if sheet.strip().casefold() == "summary":
                continue
            df = _read_station_sheet(wb, sheet)
            if df.empty:
                continue
            s = df.set_index("Date")["Rainfall Value"].sort_index()
            key = f"{_normalize_name(sheet)}|{wb.stem}"
            series_by_key[key] = s
            display_by_key[key] = sheet
            workbook_by_key[key] = wb.stem
    return series_by_key, display_by_key, workbook_by_key


def _count_missing_for_station(series: pd.Series, valid_index: pd.DatetimeIndex) -> int:
    if len(valid_index) == 0:
        return 0
    return int(series.reindex(valid_index).isna().sum())


def _load_final_mapping(final_mapped: Path) -> dict[str, str]:
    """
    Map shapefile station name -> excel station sheet name (best effort global mapping).
    """
    mapping: dict[str, str] = {}
    if not final_mapped.exists():
        return mapping
    xls = pd.ExcelFile(final_mapped)
    for sheet in xls.sheet_names:
        df = pd.read_excel(final_mapped, sheet_name=sheet)
        if "Shapefile Station Name" not in df.columns or "Excel Station Sheet Name" not in df.columns:
            continue
        for _, row in df.iterrows():
            shp = row.get("Shapefile Station Name")
            exs = row.get("Excel Station Sheet Name")
            if pd.isna(shp) or pd.isna(exs):
                continue
            shp_s = str(shp).strip()
            exs_s = str(exs).strip()
            if shp_s and exs_s:
                mapping[_normalize_name(shp_s)] = exs_s
    return mapping


def _resolve_neighbor_key(
    neighbor_name: str,
    target_group: str,
    all_keys: list[str],
    display_by_key: dict[str, str],
    final_map: dict[str, str],
) -> str | None:
    # If neighbor name is shapefile-style, map to excel sheet name first.
    mapped_name = final_map.get(_normalize_name(neighbor_name), neighbor_name)
    neighbor_candidates = _name_keys(mapped_name)

    # Prefer same target workbook first.
    same_group = [k for k in all_keys if k.endswith(f"|{target_group}")]
    all_scopes = [same_group, all_keys]

    for scope in all_scopes:
        for k in scope:
            sheet_name = display_by_key[k]
            for nk in neighbor_candidates:
                if nk in _name_keys(sheet_name):
                    return k
    return None


def _select_neighbors_for_target(
    target_station: str,
    target_group: str,
    neighbor_report: pd.DataFrame,
    all_keys: list[str],
    display_by_key: dict[str, str],
    final_map: dict[str, str],
) -> list[tuple[str, float]]:
    row = neighbor_report[neighbor_report["Target Station"].astype(str).str.strip() == target_station]
    if row.empty:
        return []
    r = row.iloc[0]

    out: list[tuple[str, float]] = []
    for i in range(1, 5):
        ncol = f"Neighbour_{i}_Name"
        dcol = f"Neighbour_{i}_Distance_from_Target_km"
        nname = r.get(ncol)
        dist = pd.to_numeric(r.get(dcol), errors="coerce")
        if pd.isna(nname) or str(nname).strip() == "":
            continue
        key = _resolve_neighbor_key(str(nname).strip(), target_group, all_keys, display_by_key, final_map)
        if key is None:
            continue
        if pd.isna(dist):
            dist = math.nan
        out.append((key, float(dist)))
    # de-duplicate preserving order
    seen = set()
    deduped: list[tuple[str, float]] = []
    for k, d in out:
        if k in seen:
            continue
        seen.add(k)
        deduped.append((k, d))
    return deduped


def _choose_best_prefix(
    target_series: pd.Series, wide: pd.DataFrame, neighbors: list[str]
) -> tuple[list[str], dict[str, float | int | None]]:
    if len(neighbors) <= 1:
        return neighbors, {
            "n_holdout_points": 0,
            "mae_mm": None,
            "rmse_mm": None,
            "holdout_r2": None,
            "mean_actual_mm": None,
        }
    candidates: list[tuple[float, float, float, int, list[str], int, float]] = []
    for k in range(2, len(neighbors) + 1):
        subset = neighbors[:k]
        X = wide[subset]
        mask = target_series.notna() & X.notna().all(axis=1)
        n_train = int(mask.sum())
        if n_train < MIN_TRAIN_ROWS:
            continue
        X_train = X[mask].values
        y_train = target_series[mask].values
        idx = np.arange(n_train)
        rng = np.random.default_rng(42)
        rng.shuffle(idx)
        n_hold = max(int(n_train * HOLDOUT_FRACTION), 10)
        hold_idx, fit_idx = idx[:n_hold], idx[n_hold:]
        scaler = StandardScaler().fit(X_train[fit_idx])
        model = _build_mlp_model().fit(scaler.transform(X_train[fit_idx]), y_train[fit_idx])
        pred = np.clip(model.predict(scaler.transform(X_train[hold_idx])), 0, None)
        mae = mean_absolute_error(y_train[hold_idx], pred)
        rmse = math.sqrt(mean_squared_error(y_train[hold_idx], pred))
        r2 = r2_score(y_train[hold_idx], pred) if len(hold_idx) > 1 else np.nan
        mean_actual = float(np.mean(y_train[hold_idx])) if len(hold_idx) > 0 else np.nan
        candidates.append((mae, rmse, -k, r2, subset, len(hold_idx), mean_actual))
    if candidates:
        candidates.sort(key=lambda x: (x[0], x[1], x[2]))
        best = candidates[0]
        return best[4], {
            "n_holdout_points": int(best[5]),
            "mae_mm": float(best[0]),
            "rmse_mm": float(best[1]),
            "holdout_r2": float(best[3]) if pd.notna(best[3]) else None,
            "mean_actual_mm": float(best[6]) if pd.notna(best[6]) else None,
        }
    return neighbors[:1], {
        "n_holdout_points": 0,
        "mae_mm": None,
        "rmse_mm": None,
        "holdout_r2": None,
        "mean_actual_mm": None,
    }


def _fit_for_subset(target_series: pd.Series, wide: pd.DataFrame, subset: list[str]):
    X = wide[subset]
    mask = target_series.notna() & X.notna().all(axis=1)
    if int(mask.sum()) < MIN_TRAIN_ROWS:
        return None
    X_train = X[mask].values
    y_train = target_series[mask].values
    scaler = StandardScaler().fit(X_train)
    model = _build_mlp_model().fit(scaler.transform(X_train), y_train)
    return scaler, model


def _mlp_fill_target(
    target_series: pd.Series,
    wide: pd.DataFrame,
    neighbor_keys: list[str],
) -> tuple[pd.Series, dict[str, float | int | None]]:
    if not neighbor_keys:
        return pd.Series(dtype=float), {
            "n_holdout_points": 0,
            "mae_mm": None,
            "rmse_mm": None,
            "holdout_r2": None,
            "mean_actual_mm": None,
        }

    neighbors, eval_info = _choose_best_prefix(target_series, wide, neighbor_keys)
    X_full = wide[neighbors]

    if target_series.notna().any():
        active_min = target_series.dropna().index.min()
        active_max = target_series.dropna().index.max()
    else:
        return pd.Series(dtype=float), eval_info

    in_active = (X_full.index >= active_min) & (X_full.index <= active_max)
    missing_target = target_series.isna() & in_active
    avail_mat = X_full.notna()
    candidate_dates = missing_target & avail_mat.any(axis=1)

    patterns: dict[frozenset[str], list[pd.Timestamp]] = {}
    for d in X_full.index[candidate_dates]:
        subset = frozenset([n for n in neighbors if avail_mat.at[d, n]])
        if subset:
            patterns.setdefault(subset, []).append(d)

    cache: dict[frozenset[str], tuple[StandardScaler, MLPRegressor]] = {}
    filled: dict[pd.Timestamp, float] = {}

    for subset_keys in sorted(patterns.keys(), key=lambda s: (-len(s), sorted(s))):
        dates = patterns[subset_keys]
        fit = cache.get(subset_keys)
        if fit is None:
            sub_cols = sorted(subset_keys)
            fit = _fit_for_subset(target_series, wide, sub_cols)
            if fit is not None:
                cache[subset_keys] = fit
        if fit is None:
            continue

        scaler, model = fit
        sub_cols = sorted(subset_keys)
        X_arr = wide.loc[dates, sub_cols].values
        if np.isnan(X_arr).any():
            continue
        pred = np.clip(model.predict(scaler.transform(X_arr)), 0, None)
        for d, v in zip(dates, pred):
            if d not in filled:
                filled[d] = float(v)
    out_series = pd.Series(filled).sort_index() if filled else pd.Series(dtype=float)
    return out_series, eval_info


def _temporal_fill(target_series: pd.Series, filled_series: pd.Series, flags: pd.Series) -> tuple[pd.Series, pd.Series]:
    s = filled_series.copy()
    candidate = target_series.isna() & flags.eq("M")
    s_interp = s.interpolate(method="linear", limit=TEMPORAL_INTERP_MAX_GAP_DAYS, limit_area="inside")
    new_fill = candidate & s_interp.notna()
    s.loc[new_fill] = s_interp.loc[new_fill]
    flags.loc[new_fill] = "T"
    return s, flags


def _fallback_aa_iw_nr(
    target_key: str,
    wide_filled: pd.DataFrame,
    flags_col: pd.Series,
    neighbor_info: list[tuple[str, float]],
) -> tuple[pd.Series, pd.Series]:
    t_series = wide_filled[target_key].copy()
    station_means = wide_filled.mean(skipna=True)
    missing_dates = list(t_series.index[flags_col.eq("M")])
    target_mean = station_means.get(target_key, np.nan)

    for d in missing_dates:
        vals = []
        vals_dist = []
        vals_nr = []
        for nk, dist in neighbor_info:
            if nk not in wide_filled.columns:
                continue
            v = wide_filled.at[d, nk]
            if pd.isna(v):
                continue
            vv = float(v)
            vals.append(vv)
            vals_dist.append((vv, dist))
            mn = station_means.get(nk, np.nan)
            if pd.notna(mn) and float(mn) > 0:
                vals_nr.append(vv / float(mn))

        fill_v = None
        fill_flag = None
        if len(vals) >= 3:
            fill_v = float(np.mean(vals))
            fill_flag = "AA"
        elif len(vals_dist) >= 2:
            wsum = 0.0
            vsum = 0.0
            for vv, dd in vals_dist:
                if pd.notna(dd) and dd > 0:
                    w = 1.0 / dd
                    wsum += w
                    vsum += w * vv
            if wsum > 0:
                fill_v = float(vsum / wsum)
                fill_flag = "IW"
        if fill_v is None and len(vals_nr) >= 1 and pd.notna(target_mean) and float(target_mean) > 0:
            fill_v = float(target_mean) * float(np.mean(vals_nr))
            fill_flag = "NR"

        if fill_v is not None:
            fill_v = max(fill_v, 0.0)
            t_series.at[d] = fill_v
            flags_col.at[d] = fill_flag

    return t_series, flags_col


def _save_error_summary_all_stations(path: Path, station_rows: list[dict[str, object]]) -> None:
    columns = [
        "target_station",
        "n_holdout_points",
        "mae_mm",
        "rmse_mm",
        "holdout_r2",
        "mean_actual_mm",
        "nmae_percent",
        "smape_percent",
        "Easy_Correct_%",
        "Easy_Wrong_%",
        "Toatl missing values",
    ]
    df = pd.DataFrame(station_rows)
    if df.empty:
        df = pd.DataFrame(columns=columns)

    for c in columns:
        if c not in df.columns:
            df[c] = np.nan
    df = df[columns].copy()

    # OVERALL row
    overall = {"target_station": "OVERALL"}
    for c in columns[1:]:
        vals = pd.to_numeric(df[c], errors="coerce").dropna()
        if len(vals) == 0:
            overall[c] = np.nan
        elif c == "Toatl missing values":
            overall[c] = int(vals.sum())
        else:
            overall[c] = float(vals.mean())

    # MEAN row
    mean_row = {"target_station": "MEAN"}
    for c in columns[1:]:
        vals = pd.to_numeric(df[c], errors="coerce").dropna()
        mean_row[c] = float(vals.mean()) if len(vals) else np.nan

    out = pd.concat([pd.DataFrame([overall]), df.sort_values("target_station"), pd.DataFrame([mean_row])], ignore_index=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Error_Summary"
    ws.cell(1, 1, "Easy_Correct_% and Easy_Wrong_% are based on NMAE% (Wrong = NMAE%, Correct = 100 - Wrong).")
    ws.cell(2, 1, "Technical metrics (MAE/RMSE/R²) are primary.")
    ws.cell(3, 1, "This summary is generated from current run metrics and final missing flags.")
    for i, c in enumerate(columns, start=1):
        ws.cell(5, i, c)
    start_row = 6
    for r_idx, (_, row) in enumerate(out.iterrows(), start=start_row):
        for c_idx, c in enumerate(columns, start=1):
            ws.cell(r_idx, c_idx, row[c])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def build_filled_values() -> None:
    print("Starting fill pipeline (MLP + temporal + AA/IW/NR)...")
    if not REQUIRED_DIR.exists():
        raise FileNotFoundError(f"Required Station Info folder not found: {REQUIRED_DIR}")
    if not NEIGHBOR_REPORT.exists():
        raise FileNotFoundError(f"Neighbor report not found: {NEIGHBOR_REPORT}")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"Input folder: {REQUIRED_DIR}")
    print(f"Output folder: {OUTPUT_DIR}")
    print(f"Neighbor report: {NEIGHBOR_REPORT}")

    neighbor_report = pd.read_excel(NEIGHBOR_REPORT)
    final_map = _load_final_mapping(FINAL_MAPPED_FILE)
    print(f"Loaded neighbor rows: {len(neighbor_report)}")
    print(f"Loaded final mapping entries: {len(final_map)}")

    series_by_key, display_by_key, workbook_by_key = _load_all_station_series(REQUIRED_DIR)
    all_keys = list(series_by_key.keys())
    valid_index_by_key: dict[str, pd.DatetimeIndex] = {
        k: pd.DatetimeIndex(s.index).drop_duplicates().sort_values() for k, s in series_by_key.items()
    }
    print(f"Loaded station series: {len(all_keys)}")

    all_dates = pd.Index([])
    for s in series_by_key.values():
        all_dates = all_dates.union(s.index)
    all_dates = all_dates.sort_values()
    print(f"Combined date index length: {len(all_dates)}")

    wide_observed = pd.DataFrame(index=all_dates)
    for k, s in series_by_key.items():
        wide_observed[k] = s.reindex(all_dates)

    wide_filled = wide_observed.copy()
    flags = wide_observed.notna().map(lambda x: "O" if x else "M")
    total_missing_initial = int(
        sum(_count_missing_for_station(wide_observed[k], valid_index_by_key[k]) for k in all_keys)
    )
    print(f"Initial missing cells across matrix: {total_missing_initial}")

    # Fill target-by-target using neighbor report.
    n_targets = len(all_keys)
    filled_targets = 0
    summary_rows: list[dict[str, object]] = []
    for idx, target_key in enumerate(all_keys, start=1):
        target_station = display_by_key[target_key]
        target_group = workbook_by_key[target_key]
        before_missing = _count_missing_for_station(wide_filled[target_key], valid_index_by_key[target_key])

        neigh_info = _select_neighbors_for_target(
            target_station=target_station,
            target_group=target_group,
            neighbor_report=neighbor_report,
            all_keys=all_keys,
            display_by_key=display_by_key,
            final_map=final_map,
        )
        neigh_keys = [k for k, _d in neigh_info if k in wide_filled.columns and k != target_key]
        if not neigh_keys:
            summary_rows.append(
                {
                    "target_station": f"{target_station} ({target_group})",
                    "n_holdout_points": 0,
                    "mae_mm": np.nan,
                    "rmse_mm": np.nan,
                    "holdout_r2": np.nan,
                    "mean_actual_mm": np.nan,
                    "nmae_percent": np.nan,
                    "smape_percent": np.nan,
                    "Easy_Correct_%": np.nan,
                    "Easy_Wrong_%": np.nan,
                    "Toatl missing values": before_missing,
                }
            )
            if idx % 20 == 0 or idx == n_targets:
                print(f"[{idx}/{n_targets}] {target_station} ({target_group}) - no usable neighbors, skipped")
            continue

        # MLP-based fill
        filled_series, eval_info = _mlp_fill_target(wide_filled[target_key], wide_filled, neigh_keys)
        if not filled_series.empty:
            wide_filled.loc[filled_series.index, target_key] = filled_series.values
            flags.loc[filled_series.index, target_key] = "F"

        # Temporal interpolation fallback
        tmp_series, tmp_flags = _temporal_fill(
            target_series=wide_observed[target_key],
            filled_series=wide_filled[target_key],
            flags=flags[target_key],
        )
        wide_filled[target_key] = tmp_series
        flags[target_key] = tmp_flags

        # AA -> IW -> NR fallback
        fb_series, fb_flags = _fallback_aa_iw_nr(target_key, wide_filled, flags[target_key], neigh_info)
        wide_filled[target_key] = fb_series
        flags[target_key] = fb_flags
        after_missing = _count_missing_for_station(wide_filled[target_key], valid_index_by_key[target_key])
        if after_missing < before_missing:
            filled_targets += 1

        mae = eval_info.get("mae_mm")
        mean_actual = eval_info.get("mean_actual_mm")
        nmae = (100.0 * float(mae) / float(mean_actual)) if (mae is not None and mean_actual not in [None, 0] and pd.notna(mean_actual)) else np.nan
        easy_wrong = float(np.clip(nmae, 0, 100)) if pd.notna(nmae) else np.nan
        easy_correct = float(100.0 - easy_wrong) if pd.notna(easy_wrong) else np.nan
        summary_rows.append(
            {
                "target_station": f"{target_station} ({target_group})",
                "n_holdout_points": eval_info.get("n_holdout_points", 0),
                "mae_mm": mae,
                "rmse_mm": eval_info.get("rmse_mm"),
                "holdout_r2": eval_info.get("holdout_r2"),
                "mean_actual_mm": mean_actual,
                "nmae_percent": nmae,
                "smape_percent": np.nan,
                "Easy_Correct_%": easy_correct,
                "Easy_Wrong_%": easy_wrong,
                "Toatl missing values": after_missing,
            }
        )

        if idx % 10 == 0 or idx == n_targets:
            print(
                f"[{idx}/{n_targets}] {target_station} ({target_group}) - "
                f"missing {before_missing} -> {after_missing} using {len(neigh_keys)} neighbors"
            )

    # Write 4 output workbooks mirroring Required Station Info.
    print("Writing output workbooks to Filled Values...")
    for wb in sorted(REQUIRED_DIR.glob("*.xlsx")):
        if wb.name.startswith("~$"):
            continue
        out_path = OUTPUT_DIR / wb.name
        print(f"  Processing workbook: {wb.name}")
        xls = pd.ExcelFile(wb)
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            for sheet in xls.sheet_names:
                if sheet.strip().casefold() == "summary":
                    # Skip old summary; can be regenerated later.
                    continue
                in_df = pd.read_excel(wb, sheet_name=sheet)
                if "Date" not in in_df.columns or "Rainfall Value" not in in_df.columns:
                    in_df.to_excel(writer, sheet_name=sheet[:31], index=False)
                    continue

                key = f"{_normalize_name(sheet)}|{wb.stem}"
                if key not in wide_filled.columns:
                    in_df.to_excel(writer, sheet_name=sheet[:31], index=False)
                    continue

                work = in_df.copy()
                dts = pd.to_datetime(work["Date"], errors="coerce")
                predicted = pd.Series(index=work.index, dtype=float)
                for i, d in dts.items():
                    if pd.isna(d):
                        continue
                    val = wide_filled.at[d, key] if d in wide_filled.index else np.nan
                    if pd.notna(val):
                        predicted.at[i] = float(val)
                # Fill only missing rainfall cells.
                rain_num = pd.to_numeric(work["Rainfall Value"], errors="coerce")
                mask_missing = rain_num.isna() & predicted.notna()
                work.loc[mask_missing, "Rainfall Value"] = predicted.loc[mask_missing].round(3)
                work.to_excel(writer, sheet_name=sheet[:31], index=False)

        print(f"Created filled workbook: {out_path}")

    total_missing_final = int(
        sum(_count_missing_for_station(wide_filled[k], valid_index_by_key[k]) for k in all_keys)
    )
    _save_error_summary_all_stations(ERROR_SUMMARY_FILE, summary_rows)
    print(f"Saved error summary workbook: {ERROR_SUMMARY_FILE}")
    print("Fill pipeline completed.")
    print(f"Targets with improved missing counts: {filled_targets}/{n_targets}")
    print(f"Total missing cells: {total_missing_initial} -> {total_missing_final}")


if __name__ == "__main__":
    build_filled_values()

