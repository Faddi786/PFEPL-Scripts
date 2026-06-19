from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font


ROOT_DIR = Path(__file__).resolve().parents[1]
SOURCE_DIR = ROOT_DIR / "Filled Values"
OUTPUT_FILE = ROOT_DIR / "Filled Values" / "Filled_Values_Merged.xlsx"
INDEX_SHEET = "Index"


def _sanitize_sheet_name(name: str) -> str:
    bad = ['\\', '/', '*', '?', ':', '[', ']']
    out = name
    for ch in bad:
        out = out.replace(ch, "_")
    return out[:31]


def _unique_sheet_name(base: str, used: set[str]) -> str:
    name = _sanitize_sheet_name(base)
    if name not in used:
        used.add(name)
        return name
    i = 2
    while True:
        suffix = f"_{i}"
        candidate = _sanitize_sheet_name(name[: 31 - len(suffix)] + suffix)
        if candidate not in used:
            used.add(candidate)
            return candidate
        i += 1


def merge_required_station_excels(source_dir: Path, output_file: Path) -> None:
    if not source_dir.exists():
        raise FileNotFoundError(f"Source folder not found: {source_dir}")

    excel_files = sorted([p for p in source_dir.glob("*.xlsx") if not p.name.startswith("~$")])
    if not excel_files:
        raise FileNotFoundError(f"No .xlsx files found in: {source_dir}")

    used_sheet_names = {INDEX_SHEET}
    index_rows: list[dict[str, str]] = []

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        # Placeholder index sheet first.
        pd.DataFrame({"Sheet Name": [], "Source Workbook": []}).to_excel(
            writer, sheet_name=INDEX_SHEET, index=False
        )

        for wb_path in excel_files:
            wb_name = wb_path.stem
            xls = pd.ExcelFile(wb_path)
            for src_sheet in xls.sheet_names:
                df = pd.read_excel(wb_path, sheet_name=src_sheet)
                merged_sheet_name = _unique_sheet_name(f"{wb_name}_{src_sheet}", used_sheet_names)
                df.to_excel(writer, sheet_name=merged_sheet_name, index=False)
                index_rows.append(
                    {
                        "Sheet Name": merged_sheet_name,
                        "Source Workbook": wb_path.name,
                    }
                )

        index_df = pd.DataFrame(index_rows)
        index_df.to_excel(writer, sheet_name=INDEX_SHEET, index=False)

    # Add clickable hyperlinks in first column of index sheet.
    wb = load_workbook(output_file)
    ws = wb[INDEX_SHEET]
    hyperlink_font = Font(color="0000EE", underline="single")
    for row in range(2, ws.max_row + 1):
        sheet_name = ws.cell(row=row, column=1).value
        if not sheet_name:
            continue
        ws.cell(row=row, column=1).hyperlink = f"#'{sheet_name}'!A1"
        ws.cell(row=row, column=1).font = hyperlink_font
    wb.save(output_file)

    print(f"Merged workbook created: {output_file}")
    print(f"Source workbooks merged: {len(excel_files)}")
    print(f"Total sheets copied: {len(index_rows)}")


if __name__ == "__main__":
    merge_required_station_excels(SOURCE_DIR, OUTPUT_FILE)

