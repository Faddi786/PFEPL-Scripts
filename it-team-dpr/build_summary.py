from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
INPUT_DIR = SCRIPT_DIR / "input"
OUTPUT_DIR = SCRIPT_DIR / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

SOURCE_FILE = str(INPUT_DIR / "it_team_dpr.xlsx")
OUTPUT_FILE = str(OUTPUT_DIR / "it_team_dpr_summary.xlsx")


def load_records(path_str: str = SOURCE_FILE):
    """Load all rows from all sheets into a flat list of records."""
    path = Path(path_str)
    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")

    wb = load_workbook(path, data_only=True)
    records = []

    for ws in wb.worksheets:
        date_label = ws.title  # use sheet name as date label

        header = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            # skip completely empty rows
            if row is None or all(cell is None for cell in row):
                continue

            if header is None:
                # assume first non-empty row is header: [Name, project name, task description, tomorrow task]
                header = [str(c).strip() if c is not None else "" for c in row]
                continue

            cells = list(row) + [None] * 4
            name_raw, project_raw, task_raw, tomorrow_raw = cells[:4]

            if name_raw is None:
                continue

            name = str(name_raw).strip()
            project = "" if project_raw is None else str(project_raw).strip()
            task = "" if task_raw is None else str(task_raw).strip()
            tomorrow = "" if tomorrow_raw is None else str(tomorrow_raw).strip()

            # skip obvious absent rows or placeholder '-'
            lower_name = name.lower()
            if "absent" in lower_name:
                continue
            if name == "-" and project == "-" and task == "-":
                continue

            records.append(
                {
                    "date": date_label,
                    "name": name,
                    "project": project,
                    "task": task,
                    "tomorrow": tomorrow,
                }
            )

    return records


def group_by_person(records):
    people = defaultdict(list)
    for rec in records:
        key = rec["name"].strip()
        people[key].append(rec)
    return people


def classify_area(text: str) -> set[str]:
    """Roughly classify a project/task string into high-level work areas."""
    if not text:
        return set()
    t = text.lower()
    areas: set[str] = set()

    if "task-bot" in t or "task bot" in t or "telegram bot" in t or "tel-bot" in t:
        areas.add("Task-bot")
    if "security" in t or "csrf" in t or "xss" in t or "waf" in t or "fail2ban" in t or "https" in t or "cors" in t:
        areas.add("Security")
    if "invoice" in t:
        areas.add("Invoice")
    if "travel" in t:
        areas.add("Travel")
    if "dgps" in t or "uv dpr" in t or "uv dgps" in t:
        areas.add("DGPS DPR")
    if "cidco" in t:
        areas.add("CIDCO")
    if "hec-hms" in t:
        areas.add("HEC-HMS")
    if "sra" in t:
        areas.add("SRA")

    # Fallback: if nothing matched but we have some text, mark as "Other"
    if not areas and t.strip():
        areas.add("Other")

    return areas


def build_summary_workbook(records, output_path: str = OUTPUT_FILE):
    people = group_by_person(records)

    wb_out = Workbook()

    # Combined sheet (all rows)
    ws_all = wb_out.active
    ws_all.title = "All Entries"
    ws_all.append(["Date (Sheet)", "Name", "Project", "Task Description", "Tomorrow Task"])
    for r in records:
        ws_all.append([r["date"], r["name"], r["project"], r["task"], r["tomorrow"]])

    # Summary by person
    ws_sum = wb_out.create_sheet(title="Summary by Person")
    ws_sum.append(
        [
            "Name",
            "Days with Entries",
            "Distinct Projects",
            "Sample Work Summary",
        ]
    )

    for name, recs in sorted(people.items(), key=lambda x: x[0].lower()):
        dates = {r["date"] for r in recs}
        projects = {r["project"] for r in recs if r["project"] and r["project"] != "-"}

        # Build a compact summary: first few task descriptions with dates
        snippets = []
        # sort by date label string for stability
        for r in sorted(recs, key=lambda r: r["date"])[:8]:
            snippets.append(f"{r['date']}: {r['task']}")

        summary_text = "\n".join(snippets)

        ws_sum.append(
            [
                name,
                len(dates),
                ", ".join(sorted(projects)),
                summary_text,
            ]
        )

    # Single manager-style sheet
    ws_mgr = wb_out.create_sheet(title="Manager Summary")
    ws_mgr.append(
        [
            "Name",
            "From Date (Sheet)",
            "To Date (Sheet)",
            "Key Work Areas (e.g. task-bot, security, invoice, travel)",
        ]
    )

    for name, recs in sorted(people.items(), key=lambda x: x[0].lower()):
        # sort dates as strings (all are in same month/year naming scheme)
        dates_sorted = sorted({r["date"] for r in recs})
        from_date = dates_sorted[0] if dates_sorted else ""
        to_date = dates_sorted[-1] if dates_sorted else ""

        areas: set[str] = set()
        for r in recs:
            areas |= classify_area(r["project"])
            areas |= classify_area(r["task"])

        ws_mgr.append(
            [
                name,
                from_date,
                to_date,
                ", ".join(sorted(areas)),
            ]
        )

    wb_out.save(output_path)
    return output_path, people


def main():
    records = load_records()
    output_path, people = build_summary_workbook(records)

    print(f"Summary workbook created: {Path(output_path).resolve()}")
    print("\nDays with entries per person:")
    for name, recs in sorted(people.items(), key=lambda x: x[0].lower()):
        dates = {r["date"] for r in recs}
        print(f"- {name}: {len(dates)} days")


if __name__ == "__main__":
    main()


