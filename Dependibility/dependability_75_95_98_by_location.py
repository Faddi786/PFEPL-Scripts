import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def sanitize_sheet_name(name: str) -> str:
    """Clean sheet name to be Excel-compatible (max 31 chars, no illegal characters)."""
    name = re.sub(r'[:\\\/\?\*\[\]]', '_', str(name))
    return name[:31]


def find_closest_row_below_target(ws, dep_col_idx, target_pct: float):
    """
    Find the highest row where Dependability <= target_pct (conservative).
    If none <= target, return the closest row overall.
    """
    closest_row = None
    closest_diff = float('inf')
    le_row = None
    le_diff = float('inf')

    for row in range(2, ws.max_row + 1):
        dep_val = ws.cell(row, dep_col_idx).value
        if not isinstance(dep_val, (int, float)) or pd.isna(dep_val):
            continue

        diff = abs(dep_val - target_pct)

        # Track overall closest
        if diff < closest_diff:
            closest_diff = diff
            closest_row = row

        # Track largest <= target (most conservative for dependability)
        if dep_val <= target_pct:
            this_diff = target_pct - dep_val
            if this_diff < le_diff:
                le_diff = this_diff
                le_row = row

    return le_row if le_row is not None else closest_row


def process_location_files(input_excel_path: str, location_sheet_names: list):
    """
    Main function:
    - Reads each location sheet from input Excel
    - Creates one output .xlsx per location
    - For each month/season column → creates dependability sheet + interpolated values
    """
    for loc_sheet in location_sheet_names:
        output_file = f"{sanitize_sheet_name(loc_sheet)}.xlsx"

        # Start fresh: delete existing output file if present
        if os.path.exists(output_file):
            os.remove(output_file)

        print(f"\nProcessing location: {loc_sheet} → {output_file}")

        # Read data for this location
        try:
            df = pd.read_excel(input_excel_path, sheet_name=loc_sheet)
        except Exception as e:
            print(f"  Error reading sheet '{loc_sheet}': {e}")
            continue

        # Identify value columns (exclude 'Year')
        value_columns = [c for c in df.columns if c.lower() != 'year' and '(mcm)' in c.lower()]

        if not value_columns:
            print(f"  No MCM columns found in {loc_sheet}")
            continue

        wb = None  # Will be created on first write

        for col_name in value_columns:
            # Prepare dependability table
            temp = df[['Year', col_name]].copy()
            temp = temp.dropna(subset=[col_name])  # drop rows with missing value
            temp = temp.sort_values(by=col_name, ascending=False).reset_index(drop=True)
            temp['Rank'] = range(1, len(temp) + 1)
            n = len(temp)
            temp['Dependability (%)'] = temp['Rank'].apply(lambda r: (r / (n + 1)) * 100)
            temp = temp.rename(columns={col_name: 'Value (MCM)'})

            # Desired column order
            temp = temp[['Year', 'Value (MCM)', 'Rank', 'Dependability (%)']]

            sheet_name = sanitize_sheet_name(col_name.replace('(MCM)', '').strip())

            # Write table
            mode = 'w' if wb is None else 'a'
            with pd.ExcelWriter(output_file, engine='openpyxl', mode=mode,
                                if_sheet_exists='replace') as writer:
                temp.to_excel(writer, sheet_name=sheet_name, index=False)

            # Load workbook to add formatting & formulas
            if wb is None:
                wb = load_workbook(output_file)
            else:
                wb = load_workbook(output_file)

            ws = wb[sheet_name]

            # Locate columns
            dep_col_idx = None
            val_col_idx = None
            for c in range(1, ws.max_column + 1):
                header = str(ws.cell(1, c).value).strip().lower()
                if 'depend' in header:
                    dep_col_idx = c
                if 'value' in header:
                    val_col_idx = c

            if not dep_col_idx or not val_col_idx:
                print(f"  Warning: Required columns missing in sheet '{sheet_name}'")
                continue

            dep_letter = get_column_letter(dep_col_idx)
            val_letter = get_column_letter(val_col_idx)

            # Find rows for each target (prefer <= target)
            r75 = find_closest_row_below_target(ws, dep_col_idx, 75.0)
            r95 = find_closest_row_below_target(ws, dep_col_idx, 95.0)
            r98 = find_closest_row_below_target(ws, dep_col_idx, 98.0)

            if not all([r75, r95, r98]):
                print(f"  Warning: Could not find suitable rows for interpolation in {sheet_name}")
                continue

            # Place results in a clean area (e.g. column F starting row 5)
            result_row_start = 5
            label_col = 'F'
            value_col = 'G'

            # Clear previous results if any (optional safety)
            for row in range(result_row_start, result_row_start + 10):
                ws.cell(row, get_column_letter(ord(label_col) - 64)).value = None
                ws.cell(row, get_column_letter(ord(value_col) - 64)).value = None

            # Write 75%
            ws.cell(result_row_start, get_column_letter(ord(label_col) - 64)).value = "75% Dependable (interpolated)"
            ws.cell(result_row_start, get_column_letter(ord(value_col) - 64)).value = (
                f"={val_letter}{r75} + ((75 - {dep_letter}{r75}) * "
                f"(({val_letter}{r75+1} - {val_letter}{r75}) / "
                f"({dep_letter}{r75+1} - {dep_letter}{r75})))"
            )

            # Write 95%
            ws.cell(result_row_start + 2, get_column_letter(ord(label_col) - 64)).value = "95% Dependable (interpolated)"
            ws.cell(result_row_start + 2, get_column_letter(ord(value_col) - 64)).value = (
                f"={val_letter}{r95} + ((95 - {dep_letter}{r95}) * "
                f"(({val_letter}{r95+1} - {val_letter}{r95}) / "
                f"({dep_letter}{r95+1} - {dep_letter}{r95})))"
            )

            # Write 98% (direct value at closest <= 98%)
            ws.cell(result_row_start + 4, get_column_letter(ord(label_col) - 64)).value = "98% Dependable"
            ws.cell(result_row_start + 4, get_column_letter(ord(value_col) - 64)).value = f"={val_letter}{r98}"

            # Formatting
            for r_offset in [0, 2, 4]:
                row_num = result_row_start + r_offset
                for col_letter in [label_col, value_col]:
                    cell = ws.cell(row_num, get_column_letter(ord(col_letter) - 64))
                    cell.font = Font(bold=(col_letter == label_col))
                    cell.alignment = Alignment(horizontal='left' if col_letter == label_col else 'right')

            print(f"  → Sheet '{sheet_name}': 75/95/98% values added (rows {r75}, {r95}, {r98})")

        # Final save
        if wb:
            wb.save(output_file)
            print(f"Finished → {output_file} created successfully.\n")


# ────────────────────────────────────────────────
#  USAGE EXAMPLE
# ────────────────────────────────────────────────

if __name__ == "__main__":
    INPUT_EXCEL = r"C:\Users\Swapnali\Desktop\Sakina Maam Work\Dependibility\Location Sheet\input.xlsx"  # ← CHANGE THIS

    LOCATIONS = [
        "Location_1",
        "Location_2",
        "Location_3",
        "Location_4",
        "Location_5",
        "Location_6",
        "Location_7"
        # add more if needed
    ]

    process_location_files(INPUT_EXCEL, LOCATIONS)