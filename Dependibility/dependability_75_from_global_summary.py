import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

def _sanitize_sheet_name(name: str) -> str:
    name = re.sub(r'[:\\\/\?\*\[\]]', '_', str(name))
    return name[:31]

def _remove_sheet_if_exists(filepath: str, sheet_name: str):
    if not os.path.exists(filepath):
        return
    wb = load_workbook(filepath)
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
        wb.save(filepath)

def _parse_number(val):
    """Convert strings like '74.5%' to float."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        s = str(val).replace('%', '').replace(',', '').strip()
        return float(s)
    except Exception:
        return None

def seventy_five_dependiblity(global_summary_excel):

    source_sheet="global_summary"
    target_pct=75.0

    # Remove old helper sheets if present
    _remove_sheet_if_exists(global_summary_excel, "Dependability")
    _remove_sheet_if_exists(global_summary_excel, "75_Percent_Dependibility")

    # STEP 1: Read base sheet and compute dependability per junction
    df = pd.read_excel(global_summary_excel, sheet_name=source_sheet)
    junction_frames = {}
    for junction in df.columns[1:]:
        temp = df[['Year', junction]].copy()
        temp = temp.sort_values(by=junction, ascending=False).reset_index(drop=True)
        temp['Rank'] = range(1, len(temp) + 1)
        n = len(temp)
        temp['Dependability'] = temp['Rank'].apply(lambda r: (r / (n + 1)) * 100)
        temp = temp.rename(columns={junction: 'Value'})
        junction_frames[_sanitize_sheet_name(junction)] = temp

    # Write junction sheets (replace existing)
    mode = 'a' if os.path.exists(global_summary_excel) else 'w'
    with pd.ExcelWriter(global_summary_excel, engine="openpyxl", mode=mode, if_sheet_exists="replace") as writer:
        for sheet_name, frame in junction_frames.items():
            frame.to_excel(writer, sheet_name=sheet_name, index=False)

    # STEP 2: Re-open workbook and find best rows, highlight, collect summary
    wb = load_workbook(global_summary_excel)
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    summary_data = []

    print("\n=== Processing Sheets ===")
    for sheet_name in junction_frames.keys():
        if sheet_name not in wb.sheetnames:
            print(f"⚠️  Skipping missing sheet: {sheet_name}")
            continue

        ws = wb[sheet_name]
        print(f"\nSheet: {sheet_name}")

        # Find header positions (robust)
        headers = {}
        for c in range(1, ws.max_column + 1):
            raw = ws.cell(1, c).value
            if raw is None:
                continue
            headers[str(raw).strip().lower()] = c

        dep_col = next((v for k, v in headers.items() if "depend" in k), None)
        year_col = next((v for k, v in headers.items() if "year" in k), None)
        val_col = next((v for k, v in headers.items() if "value" in k), None)

        print(f"  Headers found: {list(headers.keys())}")
        print(f"  Year col: {year_col}, Dependability col: {dep_col}, Value col: {val_col}")

        if not dep_col or not year_col or not val_col:
            print("  ❌ Missing required columns, skipping.")
            continue

        dep_rows = []
        for row in range(2, ws.max_row + 1):
            num = _parse_number(ws.cell(row, dep_col).value)
            if num is None:
                continue
            dep_rows.append((row, num))

        if not dep_rows:
            print("  ❌ No numeric dependability values found, skipping.")
            continue

        # Preference: largest <= target_pct; if none, smallest >= target_pct; else closest
        le_candidates = [(r, v) for r, v in dep_rows if v <= target_pct]
        if le_candidates:
            best_row, best_val = max(le_candidates, key=lambda t: t[1])
        else:
            ge_candidates = [(r, v) for r, v in dep_rows if v >= target_pct]
            if ge_candidates:
                best_row, best_val = min(ge_candidates, key=lambda t: t[1])
            else:
                best_row, best_val = min(dep_rows, key=lambda t: abs(t[1] - target_pct))

        # Get Year and Value for the matched row
        year_val = ws.cell(best_row, year_col).value
        value_val = ws.cell(best_row, val_col).value

        # Highlight dependability cell and write E3/F3 (as before)
        ws.cell(best_row, dep_col).fill = highlight_fill
        ws.cell(3, 6).value = f"{target_pct}%"
        ws.cell(3, 7).value = year_val

        print(f"  ✅ Matched dependability: {best_val:.4f}% at row {best_row}, Year: {year_val}, Value: {value_val}")

        # collect for summary (no Value column in summary as requested)
        summary_data.append((sheet_name, year_val, round(best_val, 4), value_val))

    # STEP 3: Create summary sheet using openpyxl directly (bulletproof)
    if summary_data:
        # remove if somehow exists (double-check)
        if "75_Percent_Dependibility" in wb.sheetnames:
            wb.remove(wb["75_Percent_Dependibility"])

        summary_ws = wb.create_sheet(title="75_Percent_Dependibility", index = 1)

        # Write header (Junction, Year, Dependability (%))
        headers = ["Junctions", "Year", "Dependability (%)","Value (MCM)"]
        for col_idx, h in enumerate(headers, start=1):
            cell = summary_ws.cell(row=1, column=col_idx, value=h)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Write rows (only three columns)
        # Write rows (only three columns)
        for row_idx, (junction, year_val, dep_val, value_val) in enumerate(summary_data, start=2):
            cell = summary_ws.cell(row=row_idx, column=1, value=junction)

            # ✅ Add hyperlink to the junction sheet
            sheet_target = _sanitize_sheet_name(junction)
            cell.hyperlink = f"#'{sheet_target}'!A1"
            cell.style = "Hyperlink"                 # Apply Excel hyperlink style

            # ✅ Other columns
            summary_ws.cell(row=row_idx, column=2, value=year_val)
            summary_ws.cell(row=row_idx, column=3, value=dep_val)
            summary_ws.cell(row=row_idx, column=4, value=value_val)


        # Optional: set column widths
        col_widths = [30, 12, 18]
        for i, w in enumerate(col_widths, start=1):
            summary_ws.column_dimensions[summary_ws.cell(1, i).column_letter].width = w

        print("\n📘 Summary sheet '75_Percent_Dependibility' created with", len(summary_data), "rows.")
    else:
        print("\n⚠️ No summary data collected; summary sheet will not be created.")

    # Save workbook
    wb.save(global_summary_excel)
    from excel import autofit_columns
    autofit_columns(global_summary_excel)
    print("\n✅ Finished. Workbook updated and saved.")

# # === Example usage ===
INPUT_PATH = r"C:\Users\Swapnali\Desktop\Sakina Maam Work\Dependibility\input_jun.xlsx"
seventy_five_dependiblity(INPUT_PATH)
