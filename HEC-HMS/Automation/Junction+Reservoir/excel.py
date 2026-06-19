import os
import pandas as pd
# from config import CSV_DIR, RAW_DIR, CSV_DIR
import os
import csv
import xml.etree.ElementTree as ET


def autofit_columns(filepath):
    """
    Smart auto-fit that measures only VISIBLE text in hyperlink cells.
    Perfect for index.xlsx – columns stay narrow and beautiful.
    """
    import openpyxl
    from openpyxl.utils import get_column_letter
    import time

    if not os.path.exists(filepath):
        print(f"File not found: {filepath}")
        return

    print(f"Auto-fitting columns in: {os.path.basename(filepath)}")

    for attempt in range(1, 16):
        try:
            wb = openpyxl.load_workbook(filepath)

            for ws in wb.worksheets:
                for col in ws.columns:
                    column_letter = col[0].column_letter
                    max_length = 0

                    for cell in col:
                        if cell.value is None:
                            continue

                        # === SMART TEXT EXTRACTION ===
                        if cell.hyperlink:  # This is a real hyperlink cell
                            # Use the display text if available, otherwise fallback
                            display_text = getattr(cell, "value", "")
                            if isinstance(display_text, str) and display_text.startswith('=HYPERLINK'):
                                # Try to extract the friendly name from the formula
                                try:
                                    # Extract text between the last quote and )
                                    display_text = display_text.split('"')[-2]
                                except:
                                    pass
                            visible_text = str(display_text)
                        else:
                            visible_text = str(cell.value)

                        max_length = max(max_length, len(visible_text))

                    # === BEAUTIFUL WIDTH LOGIC (especially for index.xlsx) ===
                    if "index" in filepath.lower() or "navigation" in filepath.lower():
                        # For index.xlsx: tighter caps, nicer look
                        adjusted_width = min(max_length + 3, 40)   # Max 40 is perfect
                        if column_letter == "A":   # Junction names
                            adjusted_width = min(max_length + 4, 45)
                    else:
                        # For all other files: generous but safe
                        adjusted_width = min(max_length + 3, 60)

                    ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(filepath)
            wb.close()
            print("Columns auto-fitted beautifully!")
            return

        except (PermissionError, IOError, OSError):
            print(f"  File locked by Excel... retry {attempt}/15")
            time.sleep(2.5)
        except Exception as e:
            print(f"Autofit error: {e}")
            break

    print("Could not auto-fit – close the file in Excel and retry.")

def safe_text(elem, attr):
    """Return float or text safely"""
    if elem is None:
        return ""
    val = elem.attrib.get(attr, "")
    try:
        return float(val)
    except:
        return val

def parse_results_xml(xml_path):
    tree = ET.parse(xml_path)
    root = tree.getroot()

    rows = []
    for basin in root.findall("BasinElement"):
        name = basin.attrib.get("name", "")
        btype = basin.attrib.get("type", "")
        area_elem = basin.find("DrainageArea")
        area = safe_text(area_elem, "area")

        # Find desired metrics inside <Statistics>
        stats = {m.attrib.get("displayString"): m.attrib for m in basin.findall("Statistics/StatisticMeasure")}

        # Extract key measures
        peak_discharge = stats.get("Maximum Outflow", {}).get("value", "")
        time_of_peak   = stats.get("Time of Maximum Outflow", {}).get("value", "")
        volume_m3      = stats.get("Outflow Volume", {}).get("value", "")
        volume_units   = stats.get("Outflow Volume", {}).get("units", "M3")
        vol_mm         = stats.get("Outflow Depth", {}).get("value", "")

        rows.append({
            "Hydrologic Element": name,
            "Type": btype,
            "Drainage Area (KM2)": area,
            "Peak Discharge (M3/S)": peak_discharge,
            "Time of Peak": time_of_peak,
            "Volume (MM)": vol_mm,
            "Volume (M3)": volume_m3,
            "Volume Units": volume_units
        })
    return rows

def result_to_csv(RAW_DIR,CSV_DIR):
    print("📁 [DEBUG] Entered result_to_csv()")
    print("📂 [DEBUG] RAW_DIR:", RAW_DIR)
    print("📂 [DEBUG] CSV_DIR:", CSV_DIR)
    print("🔎 [DEBUG] Listing files in RAW_DIR...")

    # List all files in RAW_DIR
    all_files = os.listdir(RAW_DIR)
    print("📄 [DEBUG] Found {} total files: {}".format(len(all_files), all_files))

    processed_count = 0
    skipped_count = 0

    for file in all_files:
        print("\n➡️ [DEBUG] Checking file:", file)

        # Skip non-results files
        if not file.lower().endswith(".results"):
            print("⏭️ [DEBUG] Skipping (not a .results file):", file)
            skipped_count += 1
            continue

        path = os.path.join(RAW_DIR, file)
        print("📁 [DEBUG] Full file path:", path)

        try:
            # Parse the XML results file
            print("🔄 [DEBUG] Parsing XML for:", file)
            rows = parse_results_xml(path)
            print("📊 [DEBUG] parse_results_xml() returned {} rows".format(len(rows)))

            if not rows:
                print("⚠️ [DEBUG] No BasinElements found in:", file)
                continue

            # Build output CSV path
            out_csv = os.path.join(CSV_DIR, file.replace(".results", ".csv"))
            print("📝 [DEBUG] Writing parsed data to CSV:", out_csv)

            # Write CSV
            with open(out_csv, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
                print("📑 [DEBUG] CSV Header:", list(rows[0].keys()))
                writer.writeheader()
                writer.writerows(rows)

            processed_count += 1
            print("✅ [SUCCESS] Parsed {} → {} ({} elements)".format(file, out_csv, len(rows)))

        except Exception as e:
            print("❌ [ERROR] Failed {}: {}".format(file, e))

    print("\n📊 [SUMMARY]")
    print("✅ Processed .results files:", processed_count)
    print("⏭️ Skipped non-results files:", skipped_count)
    print("📁 Output CSV folder:", CSV_DIR)
    print("🎉 All done! CSVs available in:", CSV_DIR)




def csvs_transposed(CSV_DIR, global_summary_excel):
    print("we are inside csv transposed which is to be run later for result to csv")

    final_df = pd.DataFrame()
    processed_files = 0  # counter

    print(f"🔍 Scanning folder: {CSV_DIR}\n")

    for filename in os.listdir(CSV_DIR):
        if filename.endswith('.csv'):
            file_path = os.path.join(CSV_DIR, filename)
            print(f"📂 Processing file: {filename} ...")

            try:
                # Read CSV
                df = pd.read_csv(file_path)

                # ✅ Make a proper copy to avoid SettingWithCopyWarning
                # df_filtered = df[df['Type'].isin(['Junction', 'Sink'])].copy()
                # df_filtered = df[df['Type'].str.lower() == 'reservoir'].copy()
                df_filtered = df[df['Type'].str.lower().isin(['junction', 'reservoir'])].copy()
                # df_filtered = df.copy()


                # Divide Volume (M3) by 1000
                df_filtered.loc[:, 'Volume (M3)'] = ((df_filtered['Volume (M3)']) / 1000)*0.001

                # Extract year from filename (e.g., RUN_Run_2_2008.csv → 2008)
                year = filename.split('_')[-1].split('.')[0]
                df_filtered.loc[:, 'Year'] = int(year)

                # Pivot the data
                df_pivot = df_filtered.pivot(
                    index='Year', 
                    columns='Hydrologic Element', 
                    values='Volume (M3)'
                )

                # Append to final DataFrame
                final_df = pd.concat([final_df, df_pivot], axis=0)

                processed_files += 1
                print(f"✅ Successfully processed: {filename}")

            except Exception as e:
                print(f"❌ Error processing {filename}: {e}")

    # After all files processed
    if processed_files > 0:
        final_df.reset_index(inplace=True)
        output_file = global_summary_excel
        final_df.to_excel(output_file, index=False, sheet_name = "global_summary")
        print(f"\n💾 All done! Saved combined results to '{output_file}'")
        print(f"📊 Total files processed: {processed_files}")
        
        # ✅ Return list of headers excluding "Year"
        # headers = [col for col in final_df.columns if col != "Year"]
        headers = [col for col in final_df.columns]
        return headers
    
    else:
        print("⚠️ No CSV files were found or processed in this folder.")
import os
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

import os
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

def build_navigation_index(index_path, global_summary_path, timeseries_path, junctions):
    """
    Enhanced Index with 5th column: Detailed Time Series → links to Time Series/junction.xlsx
    """
    global_summary_file = os.path.basename(global_summary_path)
    timeseries_file = os.path.basename(timeseries_path)
    ts_folder_name = "Time Series"  # Must match what timeseries_full.py creates

    # Load global summary to map junction → column
    wb_global = load_workbook(global_summary_path, data_only=True)
    ws_global = wb_global["global_summary"]
    headers = [cell.value for cell in ws_global[1]]
    col_map = {}
    for junc in junctions:
        for idx, h in enumerate(headers, 1):
            if h and str(h).strip().lower() == junc.strip().lower():
                col_map[junc] = get_column_letter(idx)
                break
    wb_global.close()

    # Build dependability row map (for 75% sheet)
    dep_map = {}
    if os.path.exists(global_summary_path):
        dep_wb = load_workbook(global_summary_path, read_only=True, data_only=True)
        if "75_Percent_Dependibility" in dep_wb.sheetnames:
            ws = dep_wb["75_Percent_Dependibility"]
            for row in ws.iter_rows(min_row=2, max_col=1):
                cell = row[0]
                if cell.value:
                    dep_map[str(cell.value).strip().lower()] = cell.row
        dep_wb.close()

    # Check which time series sheets exist (for 75% file)
    wb_time = load_workbook(timeseries_path, data_only=True)
    timeseries_sheets = [s.lower() for s in wb_time.sheetnames]
    wb_time.close()

    # Output directory to find the "Time Series" folder
    output_dir = os.path.dirname(index_path)
    ts_folder_path = os.path.join(output_dir, ts_folder_name)

    records = []
    for junc in junctions:
        safe_junc = junc.replace("'", "''")
        junc_lower = junc.lower()
        junc_key = junc.strip().lower()

        # 1. Global Summary link
        col_letter = col_map.get(junc)
        if col_letter:
            global_link = f'=HYPERLINK("[{global_summary_file}]global_summary!{col_letter}1","Global Summary")'
        else:
            global_link = "Not Found"

        # 2. 75% Dependability link
        row_num = dep_map.get(junc_key)
        if row_num:
            dep_link = f'=HYPERLINK("[{global_summary_file}]75_Percent_Dependibility!A{row_num}","75% Dependability")'
        else:
            dep_link = "Not in 75% Sheet"

        # 3. Original 75% Time Series link
        if junc_lower == "year":
            time_link = f'=HYPERLINK("[{timeseries_file}]Index!A1","Time Series (75%)")' if "index" in timeseries_sheets else "Index Missing"
        elif junc_lower in timeseries_sheets:
            time_link = f'=HYPERLINK("[{timeseries_file}]{safe_junc}!A1","Time Series (75%)")'
        else:
            time_link = "Not Found"

        detailed_file = f"{junc}.xlsx"
        detailed_path = os.path.join(os.path.dirname(timeseries_path), "all_years_timeseries", detailed_file)
        
        if os.path.exists(detailed_path):
            detailed_link = f'=HYPERLINK("[all_years_timeseries\\{detailed_file}]Index!A1","Detailed Time Series")'
        else:
            detailed_link = "Generating..."

        records.append({
            "Junction Name": junc,
            "Global Summary": global_link,
            "Dependibility": dep_link,
            "Time Series (75%)": time_link,
            "Detailed Time Series": detailed_link
        })

    # Write to Index.xlsx
    df = pd.DataFrame(records)
    with pd.ExcelWriter(index_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Index", index=False)

    # Final polish
    autofit_columns(index_path)
    print(f"Enhanced Index.xlsx created with Detailed Time Series links!")
    print(f"   → {index_path}")


    # 🧭 Save to Index.xlsx
    df = pd.DataFrame(records)
    with pd.ExcelWriter(index_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Index", index=False)
    autofit_columns(index_path)

    print(f"✅ Navigation Index created successfully: {index_path}")