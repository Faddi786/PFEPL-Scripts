# -*- coding: utf-8 -*-
"""
HEC-DSS → Excel (Filtered by Junctions & Years + Pivoted Format + Smart Matching + DSS Listing)
✅ Reads dependibility_list.xlsx (Junction + Year)
✅ Smart match: handles spaces, hyphens, underscores, partial text, case-insensitivity
✅ Prints all DSS junction names after match report for verification
✅ One DSS open/close only
✅ One sheet per element (merged across months)
✅ Hyperlinked Index sheet
✅ Includes Date, Hour, and pivoted FLOW columns
Author: PFEPL Automation (final with DSS list)
"""

import os
import pandas as pd
from pydsstools.heclib.dss import HecDss
# from config import output_dss_file, global_summary_excel, DEPEND_SHEET, timeseries_excel

# === HELPERS ===
def normalize(name):
    """Remove spaces, hyphens, underscores, and lowercase everything."""
    return str(name).strip().lower().replace(" ", "").replace("-", "").replace("_", "")


def extract_parts_from_path(path):
    parts = path.strip("/").split("/")
    element = parts[0] if len(parts) > 0 else "Unknown"
    series_type = parts[1] if len(parts) > 1 else "UnknownSeries"
    run_info = parts[-1] if len(parts) > 0 else ""
    return element, series_type, run_info


def load_filters(global_summary_excel):
    """Load Excel junctions and years."""
    if not os.path.exists(global_summary_excel):
        raise FileNotFoundError(f"Dependibility file not found: {global_summary_excel}")

    df = pd.read_excel(global_summary_excel, sheet_name="75_Percent_Dependibility", engine="openpyxl")
    if "Junctions" not in df.columns or "Year" not in df.columns:
        raise ValueError(f"Sheet '75_Percent_Dependibility' must have columns 'Junctions' and 'Year'")

    df = df.dropna(subset=["Junctions", "Year"])
    df["Junctions"] = df["Junctions"].astype(str)
    df["Year"] = df["Year"].astype(int)

    return df


def best_match(excel_name, dss_dict):
    """Find the best matching DSS junction for a given Excel junction name."""
    excel_norm = normalize(excel_name)

    exact_matches = []
    partial_matches = []

    for dss_norm, dss_name in dss_dict.items():
        if excel_norm == dss_norm:  # ✅ exact match wins immediately
            exact_matches.append((dss_name, dss_norm))
        elif excel_norm in dss_norm or dss_norm in excel_norm:
            # store length difference as a measure of closeness
            score = abs(len(excel_norm) - len(dss_norm))
            partial_matches.append((dss_name, dss_norm, score))

    if exact_matches:
        return exact_matches[0]  # exact match priority

    if partial_matches:
        # choose partial with the smallest length difference (closest match)
        best = min(partial_matches, key=lambda x: x[2])
        return best[0], best[1]

    return None, None



# === MAIN ===
def timeseries(global_summary_excel,output_dss_file, timeseries_excel, for_all_years):
    excel_df = load_filters(global_summary_excel)

    print("📂 Opening DSS file once...")
    dss = HecDss.Open(output_dss_file)
    all_paths = dss.getPathnameList("")
    print(f"✅ Total paths found: {len(all_paths)}")

    # Collect all DSS elements (unique normalized form)
    dss_elements = {}
    for p in all_paths:
        element_name, _, _ = extract_parts_from_path(p)
        norm = normalize(element_name)
        if norm not in dss_elements:
            dss_elements[norm] = element_name

    # --- Report: Excel → DSS → Year ---
    print("\n========== FILTER MATCH REPORT ==========")
    excel_df["Matched_DSS"] = ""
    excel_df["Norm_Key"] = ""

    for i, row in excel_df.iterrows():
        excel_name = row["Junctions"]
        year = int(row["Year"])
        match_name, match_key = best_match(excel_name, dss_elements)
        if match_name:
            excel_df.at[i, "Matched_DSS"] = match_name
            excel_df.at[i, "Norm_Key"] = match_key
            print(f"{excel_name:<30} →  {match_name:<30} →  {year}")
        else:
            excel_df.at[i, "Matched_DSS"] = "❌ Not Found in DSS"
            print(f"{excel_name:<30} →  ❌ Not Found in DSS             →  {year}")
    print("==========================================\n")

    print("==========================================\n")

    # ✅ Check if any junction was NOT found
    unmatched = excel_df[excel_df["Matched_DSS"] == "❌ Not Found in DSS"]

    if not unmatched.empty:
        print("⚠️ WARNING: Some junctions could not be matched with DSS paths!")
        print(unmatched[["Junctions", "Year"]].to_string(index=False))
        print("\n📜 All unique junction names found in DSS (for reference):")
        print("----------------------------------------------------------")
        for i, (norm_key, original) in enumerate(sorted(dss_elements.items()), start=1):
            print(f"{i:3}. {original:<40} | normalized: {norm_key}")
        print("----------------------------------------------------------")
        print(f"Total unique DSS junctions listed: {len(dss_elements)}\n")

        # # Pause for user confirmation only if there are unmatched junctions
        # proceed = input("❓ Proceed with data extraction anyway? (y/n): ").strip().lower()
        # if proceed != "y":
        #     print("❌ Operation cancelled.")
        #     dss.close()
        #     return
    else:
        print("✅ All junction names matched successfully — proceeding automatically...\n")


    # --- Filter relevant DSS paths ---
    time_series_paths = [
        p for p in all_paths if "RUN:" in p.upper() and ("FLOW" in p.upper() or "FLOW-" in p.upper())
    ]
    print(f"📊 Time series selected: {len(time_series_paths)}")

    element_data = {}
    for p in time_series_paths:
        try:
            ts = dss.read_ts(p, trim_missing=True)
            if ts is None:
                continue

            element_name, series_type, run_info = extract_parts_from_path(p)
            norm_name = normalize(element_name)

            # --- ✅ Filter based on Excel even when for_all_years = True ---
            match_row = excel_df.loc[excel_df["Norm_Key"] == norm_name]
            if match_row.empty:
                # Element not listed in Excel, skip it
                continue

            # --- Select year only if for_all_years=False ---
            if not for_all_years:
                year = int(match_row["Year"].values[0])
            else:
                year = None  # all years allowed

            df = pd.DataFrame({
                "DateTime": pd.to_datetime(ts.pytimes),
                "Value": ts.values,
                "SeriesType": series_type,
                "Run": run_info,
                "DSS Path": p
            })

            # --- Apply year filter only if needed ---
            if not for_all_years and year is not None:
                start = pd.Timestamp(f"{year}-01-01")
                end = pd.Timestamp(f"{year}-12-31 23:59:59")
                df = df[(df["DateTime"] >= start) & (df["DateTime"] <= end)]

            if df.empty:
                continue

            dss_display_name = dss_elements[norm_name]
            element_data.setdefault(dss_display_name, []).append(df)

            if not for_all_years:
                print(f"✅ {dss_display_name} | {series_type} ({len(df)} records, year {year})")
            else:
                years_present = sorted(set(pd.to_datetime(df["DateTime"]).dt.year))
                print(f"✅ {dss_display_name} | {series_type} (All years: {years_present})")

        except Exception as e:
            print(f"❌ Error reading {p}: {e}")


    dss.close()
    print("📁 DSS file closed.\n")

    # --- Pivot and Merge ---
    merged_data = {}
    for name, dfs in element_data.items():
        combined_df = pd.concat(dfs, ignore_index=True)
        combined_df["DateTime"] = pd.to_datetime(combined_df["DateTime"])

        pivot_df = combined_df.pivot_table(
            index="DateTime",
            columns="SeriesType",
            values="Value",
            aggfunc="first"
        ).reset_index()

        pivot_df["Date"] = pivot_df["DateTime"].dt.date
        pivot_df["Hour"] = pivot_df["DateTime"].dt.strftime("%H:%M:%S")
        pivot_df = pivot_df.sort_values("DateTime").reset_index(drop=True)

        series_cols = [c for c in pivot_df.columns if c not in ("DateTime", "Date", "Hour")]
        ordered_cols = ["DateTime", "Date", "Hour"] + series_cols
        pivot_df = pivot_df[ordered_cols]

        merged_data[name] = pivot_df
        print(f"🔀 Merged & pivoted: {name} ({pivot_df.shape[0]} rows, {pivot_df.shape[1]} columns)")

    # --- Write to Excel ---
    # --- Write to Excel --- and below):

    # --- Write Output Depending on Mode ---
    print("\n📤 Writing output...")
    if for_all_years:
        import os
        os.makedirs(os.path.join(os.path.dirname(timeseries_excel), "all_years_timeseries"), exist_ok=True)
        all_years_dir = os.path.join(os.path.dirname(timeseries_excel), "all_years_timeseries")

        for name, dfs in element_data.items():
            safe_name = name[:31].replace(":", "").replace("/", "_").replace("\\", "_")
            element_path = os.path.join(all_years_dir, f"{safe_name}.xlsx")
            print(f"📘 Creating workbook for element: {name}")

            with pd.ExcelWriter(element_path, engine="openpyxl") as writer:
                index_rows = []

                # Group all dataframes by year
                df_all = pd.concat(dfs, ignore_index=True)
                df_all["Year"] = pd.to_datetime(df_all["DateTime"]).dt.year

                for y, df_y in df_all.groupby("Year"):
                    pivot_df = df_y.pivot_table(
                        index="DateTime",
                        columns="SeriesType",
                        values="Value",
                        aggfunc="first"
                    ).reset_index()
                    pivot_df["Date"] = pivot_df["DateTime"].dt.date
                    pivot_df["Hour"] = pivot_df["DateTime"].dt.strftime("%H:%M:%S")
                    series_cols = [c for c in pivot_df.columns if c not in ("DateTime", "Date", "Hour")]
                    pivot_df = pivot_df[["DateTime", "Date", "Hour"] + series_cols]

                    # Round numeric columns
                    for col in pivot_df.columns[3:8]:
                        if col in pivot_df.columns:
                            pivot_df[col] = pd.to_numeric(pivot_df[col], errors="coerce").round(1)

                    pivot_df.to_excel(writer, sheet_name=str(y), index=False)
                    index_rows.append({"S.No": len(index_rows)+1, "Year": f'=HYPERLINK("#\'{y}\'!A1","{y}")'})

                # Write index sheet
                pd.DataFrame(index_rows).to_excel(writer, sheet_name="Index", index=False)

            # Post-formatting (autofit + 1 decimal)
            from openpyxl import load_workbook
            wb = load_workbook(element_path)
            for ws in wb.sheetnames:
                if ws == "Index": continue
                ws = wb[ws]
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=8):
                    for cell in row:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = "0.0"
                for col_cells in ws.columns:
                    length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
                    ws.column_dimensions[col_cells[0].column_letter].width = length + 2
            wb.save(element_path)
            from excel import autofit_columns
            autofit_columns(element_path)
            print(f"✅ Saved: {element_path}")

        print("\n🎉 Done! All-year workbooks saved in:", all_years_dir)

    else:
            
        print("\n📤 Writing to Excel:", timeseries_excel)
        with pd.ExcelWriter(timeseries_excel, engine="openpyxl") as writer:
            index_rows = []

            # 1️⃣ Write an empty Index sheet first
            pd.DataFrame(columns=["S.No", "Element Name"]).to_excel(writer, sheet_name="Index", index=False)

            # 2️⃣ Write each data sheet
            for idx, (name, df) in enumerate(merged_data.items(), start=1):
                safe_name = name[:31].replace(":", "").replace("/", "_").replace("\\", "_")

                # Force numeric cols D–H (3:8) to 1 decimal value
                for col in df.columns[3:8]:
                    if col in df.columns:
                        df[col] = pd.to_numeric(df[col], errors="coerce").round(1)

                df.to_excel(writer, sheet_name=safe_name, index=False)

                index_rows.append({
                    "S.No": idx,
                    "Element Name": f'=HYPERLINK("#\'{safe_name}\'!A1","{name}")'
                    })

            # 3️⃣ Write Index content last (overwrite empty one)
            pd.DataFrame(index_rows).to_excel(writer, sheet_name="Index", index=False)

        # 4️⃣ Post-process workbook with openpyxl for formatting
        from openpyxl import load_workbook
        wb = load_workbook(timeseries_excel)

        # Format numeric cells (D–H) to 1 decimal and autofit columns
        for ws_name in wb.sheetnames:
            ws = wb[ws_name]

            # ✅ Apply number format for D–H
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=8):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "0.0"  # exactly one decimal place

            # ✅ Autofit columns
            for col_cells in ws.columns:
                length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
                ws.column_dimensions[col_cells[0].column_letter].width = length + 2

        wb.save(timeseries_excel)
        from excel import autofit_columns
        autofit_columns(timeseries_excel)


        print("\n🎉 Done! Excel written successfully.")
        print("📁 Output:", timeseries_excel)
        print("==========================================")


# if __name__ == "__main__":
#     main()

