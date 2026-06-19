
import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook.properties import CalcProperties
import os
from datetime import datetime
import win32com.client as win32

timestamp = datetime.now().strftime("(%d-%b-%Y)_(%I_%M_%S %p)")

# ---------------------------
# CONFIGURATIONS
# ---------------------------
DATA_FILE = "Dam_Data.xlsx"
TEMPLATE_FILE = "Dam_Template.xlsx"
OUTPUT_FOLDER = f"output_Dam_{timestamp}"

TARGET_SHEET = "Recapitulation Sheet"      # Recapitulation Sheet <-- change if needed
TARGET_CELL = "E14"       # C20 <-- change if needed

PLACEHOLDER_START = "{{"
PLACEHOLDER_END = "}}"

os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# Load data
# df = pd.read_excel(DATA_FILE)
df = pd.read_excel(DATA_FILE, header=3)



# ---------------------------------------------------------------
# FUNCTION: Replace placeholders & preserve numeric values
# ---------------------------------------------------------------
def replace_placeholders_in_sheet(ws, mapping):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                original_text = cell.value
                text = original_text

                for key, value in mapping.items():
                    placeholder = f"{PLACEHOLDER_START}{key}{PLACEHOLDER_END}"

                    if placeholder in text:

                        # Numeric placeholder ONLY
                        if text.strip() == placeholder and isinstance(value, (int, float)):
                            cell.value = value
                            break

                        # Mixed text + placeholder
                        text = text.replace(placeholder, str(value))

                else:
                    cell.value = text


def autofit_columns(filepath, min_width=10):
    """
    Smart auto-fit with minimum width control.
    - Keeps your existing hyperlink-smart autofit
    - Ensures no column becomes thinner than `min_width`
    """

    import openpyxl
    from openpyxl.utils import get_column_letter
    import time
    import os

    if not os.path.exists(filepath):
        print(f"File not found: {filepath}")
        return

    print(f"Auto-fitting columns....")
    # in: {os.path.basename(filepath)} (min width = {min_width})

    for attempt in range(1, 16):
        try:
            wb = openpyxl.load_workbook(filepath)

            for ws in wb.worksheets:
                for col in ws.columns:
                    column_letter = col[0].column_letter
                    max_length = 0

                    for cell in col:
                        if cell.value is None:
                            continue

                        # === SMART TEXT EXTRACTION ===
                        if cell.hyperlink:
                            display_text = getattr(cell, "value", "")
                            if isinstance(display_text, str) and display_text.startswith('=HYPERLINK'):
                                try:
                                    display_text = display_text.split('"')[-2]
                                except:
                                    pass
                            visible_text = str(display_text)
                        else:
                            visible_text = str(cell.value)

                        max_length = max(max_length, len(visible_text))

                    # === BEAUTIFUL WIDTH LOGIC (based on file type) ===
                    if "index" in filepath.lower() or "navigation" in filepath.lower():
                        adjusted_width = min(max_length + 3, 40)
                        if column_letter == "A":
                            adjusted_width = min(max_length + 4, 45)
                    else:
                        adjusted_width = min(max_length + 3, 60)

                    # === APPLY MINIMUM WIDTH ===
                    if adjusted_width < min_width:
                        adjusted_width = min_width

                    ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(filepath)
            wb.close()
            print("Columns auto-fitted!")
            # beautifully with min-width applied
            return

        except (PermissionError, IOError, OSError):
            print(f"  File locked by Excel... retry {attempt}/15")
            time.sleep(2.5)

        except Exception as e:
            print(f"Autofit error: {e}")
            break

    print("Could not auto-fit – close the file in Excel and retry.")


# ---------------------------------------------------------------
# STEP 1: Generate all filled Excel files
# ---------------------------------------------------------------
generated_files = []

for index, row in df.iterrows():

    mapping = {col: row[col] for col in df.columns}

    wb = load_workbook(TEMPLATE_FILE)

    for sheet in wb.sheetnames:
        replace_placeholders_in_sheet(wb[sheet], mapping)

    # Ensure formula recalc on next open
    try:
        if wb.calc_properties is not None:
            wb.calc_properties.fullCalcOnLoad = True
        else:
            wb.calc_properties = CalcProperties(fullCalcOnLoad=True)
    except:
        pass

    identifier_1 = row[df.columns[0]]
    identifier_2 = row[df.columns[4]]
    output_file = f"{identifier_1}.{identifier_2}.xlsx"
    output_path = os.path.join(OUTPUT_FOLDER, output_file)

    from openpyxl.drawing.image import Image
    
    img = Image("emblem.png")  # convert WMF → PNG once & save as logo.png
    ws = wb["Cover"]     # sheet where image goes
    ws.add_image(img, "E4")  # position where image should appear


    wb.save(output_path)
    generated_files.append(output_path)

    print(f"Generated: {output_path}")

print("All Excel documents generated successfully!")


# ---------------------------------------------------------------
# STEP 2: Use Excel COM to force calculation in each file
# ---------------------------------------------------------------
print("\nForcing Excel to calculate formulas...")

excel = win32.gencache.EnsureDispatch("Excel.Application")
excel.Visible = False
excel.DisplayAlerts = False
excel.AskToUpdateLinks = False 

for file_path in generated_files:
    try:
        wb = excel.Workbooks.Open(os.path.abspath(file_path))
        wb.Save()     # Excel recalculates formulas here
        wb.Close()
        print(f"Calculated formulas in: {file_path}")
    except Exception as e:
        print(f"Error recalculating {file_path}: {e}")

excel.Quit()
print("Excel recalculation completed!\n")

# wb.calculation.fullCalcOnLoad = True



# ---------------------------------------------------------------
# STEP 3: Create data_copy.xlsx
# ---------------------------------------------------------------
data_copy_path = os.path.join(OUTPUT_FOLDER, "0.data_copy.xlsx")
df_copy = df.copy()
df_copy["Amount"] = ""     # new column

# Column header rename mapping
COLUMN_RENAME_MAP = {
    "L": "Length",
    "I_P": "Intensity Percentage",
    "I_F": "Intensity Fraction",
    "BM_L": "Benchmark Length",
    "B": "Breadth",
    "N_B": "No. Of Bores",
    "B_R": "Bores in River",
    "B_B": "Bores on Bank",
    "O_R_D": "Overburden River Depth",
    "O_B_D": "Overburden Bank Depth",
    "H_R_D": "Hardrock River Depth",
    "H_B_D": "Hardrock Bank Depth",
    "EE_1": "Executive Engineer_1",
    "EE_2": "Executive Engineer_2",
    "SDE_1": "Sub Divisional Engineer_1",
    "SDE_2": "Sub Divisional Engineer_2",
}
df_copy.rename(columns=COLUMN_RENAME_MAP, inplace=True)
# print("Column headers renamed successfully!")

print("Created data_copy.xlsx")


# ---------------------------------------------------------------
# STEP 4: Read the computed Amount from each Excel file
# ---------------------------------------------------------------
for i, file_path in enumerate(generated_files):
    try:
        wb = load_workbook(file_path, data_only=True)
        if TARGET_SHEET not in wb.sheetnames:
            raise Exception(f"Worksheet '{TARGET_SHEET}' not found")

        ws = wb[TARGET_SHEET]
        value = ws[TARGET_CELL].value

        df_copy.loc[i, "Amount"] = value
        print(f"Extracted Amount from {file_path}: {value}")

    except Exception as e:
        print(f"Error reading {file_path}: {e}")
        df_copy.loc[i, "Amount"] = "ERROR"


# df_copy.to_excel(data_copy_path, index=False)
# print(f"\nUpdated data sheet saved at: {data_copy_path}")


# ---------------------------------------------------------------
# STEP 5: Write df_copy to Excel Starting at Row 4
# ---------------------------------------------------------------

# Temporary file
temp_path = os.path.join(OUTPUT_FOLDER, "_temp_data_copy.xlsx")

# Write normally first
df_copy.to_excel(temp_path, index=False)

# Now reposition rows using openpyxl
wb = load_workbook(temp_path)
ws = wb.active

# # Insert 3 blank rows ABOVE row 1 → moves header to Row 4
# ws.insert_rows(1, amount=3)

# # ---------------------------------------------------------------
# # COPY rows 1–3 from original data file to data_copy
# # ---------------------------------------------------------------
# orig_wb = load_workbook(DATA_FILE)
# orig_ws = orig_wb.active

# # copy values for rows 1–3
# for r in range(1, 4):
#     for c in range(1, ws.max_column + 1):
#         ws.cell(row=r, column=c).value = orig_ws.cell(row=r, column=c).value

# Save final correct file
wb.save(data_copy_path)

# Remove temporary file
os.remove(temp_path)

print(f"\nUpdated data sheet saved at: {data_copy_path}")
autofit_columns(data_copy_path, min_width=12) 

