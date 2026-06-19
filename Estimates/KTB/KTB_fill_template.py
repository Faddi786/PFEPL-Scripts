import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook.properties import CalcProperties
import os
from datetime import datetime
import time
import win32com.client as win32
from openpyxl.drawing.image import Image

# ---------------------------------------------------------------
# TIMESTAMP
# ---------------------------------------------------------------
timestamp = datetime.now().strftime("(%d-%b-%Y)_(%I_%M_%S %p)")

# ---------------------------------------------------------------
# CONFIGURATIONS
# ---------------------------------------------------------------
DATA_FILE = "KTB_Data_Latest.xlsx"
TEMPLATE_FILE = "KTB_Template.xlsx"
OUTPUT_FOLDER = f"output_KTB_data_1-25_{timestamp}"

TARGET_SHEET = "Recapitulation Sheet"
TARGET_CELL = "C20"

PLACEHOLDER_START = "{{"
PLACEHOLDER_END = "}}"

os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# ---------------------------------------------------------------
# LOAD DATA
# ---------------------------------------------------------------
df = pd.read_excel(DATA_FILE, header=3)

# ---------------------------------------------------------------
# PLACEHOLDER REPLACEMENT
# ---------------------------------------------------------------
def replace_placeholders_in_sheet(ws, mapping):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                text = cell.value
                for key, value in mapping.items():
                    placeholder = f"{PLACEHOLDER_START}{key}{PLACEHOLDER_END}"

                    if placeholder in text:
                        if text.strip() == placeholder and isinstance(value, (int, float)):
                            cell.value = value
                            break
                        text = text.replace(placeholder, str(value))
                else:
                    cell.value = text

# ---------------------------------------------------------------
# AUTOFIT COLUMNS
# ---------------------------------------------------------------
def autofit_columns(filepath, min_width=10):
    import openpyxl
    from openpyxl.utils import get_column_letter

    for attempt in range(1, 16):
        try:
            wb = openpyxl.load_workbook(filepath)

            for ws in wb.worksheets:
                for col in ws.columns:
                    column_letter = col[0].column_letter
                    max_length = 0

                    for cell in col:
                        if cell.value is None:
                            continue
                        visible_text = str(cell.value)
                        max_length = max(max_length, len(visible_text))

                    adjusted_width = min(max_length + 3, 60)
                    if adjusted_width < min_width:
                        adjusted_width = min_width

                    ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(filepath)
            wb.close()
            return

        except (PermissionError, IOError, OSError):
            time.sleep(2)

# ---------------------------------------------------------------
# STEP 1: GENERATE FILLED EXCEL FILES
# ---------------------------------------------------------------
generated_files = []

for index, row in df.iterrows():
    mapping = {col: row[col] for col in df.columns}
    wb = load_workbook(TEMPLATE_FILE)

    for sheet in wb.sheetnames:
        replace_placeholders_in_sheet(wb[sheet], mapping)

    # Ensure formulas recalc in Excel
    # if wb.calc_properties:
    #     wb.calc_properties.fullCalcOnLoad = True
    # else:
    #     wb.calc_properties = CalcProperties(fullCalcOnLoad=True)

    # Ensure formulas recalc on Excel open (safe for all versions)
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass


    identifier_1 = row[df.columns[0]]
    identifier_2 = row[df.columns[2]]
    output_file = f"{identifier_1}.{identifier_2}.xlsx"
    output_path = os.path.join(OUTPUT_FOLDER, output_file)

    img = Image("emblem.png")
    wb["cover"].add_image(img, "E4")

    wb.save(output_path)
    generated_files.append(output_path)

    print(f"Generated: {output_path}")

print("All Excel documents generated successfully!")

# ---------------------------------------------------------------
# STEP 2: FORCE EXCEL CALCULATION (PYTHON 3.14 SAFE)
# ---------------------------------------------------------------
print("\nForcing Excel to calculate formulas...")

win32.gencache.is_readonly = True

excel = win32.DispatchEx("Excel.Application")
excel.Visible = False
excel.DisplayAlerts = False
excel.AskToUpdateLinks = False

for file_path in generated_files:
    abs_path = os.path.abspath(file_path)
    wb = excel.Workbooks.Open(abs_path, UpdateLinks=0)

    excel.CalculateFullRebuild()

    while excel.CalculationState != 0:  # xlDone
        time.sleep(0.2)

    wb.Save()
    wb.Close()

excel.Quit()
print("Excel recalculation completed!\n")

# ---------------------------------------------------------------
# STEP 3: CREATE data_copy.xlsx
# ---------------------------------------------------------------
data_copy_path = os.path.join(OUTPUT_FOLDER, "0.data_copy.xlsx")
df_copy = df.copy()
df_copy["Amount"] = ""

COLUMN_RENAME_MAP = {
    "L": "Length",
    "I_P": "Intensity Percentage",
    "I_F": "Intensity Fraction",
    "BM_L": "Benchmark Length",
    "B": "Breadth",
    "N_B": "No. Of Bores",
    "B_R": "Bores in River",
    "B_B": "Bores on Bank",
    "O_R_D": "Overburden River Depth",
    "O_B_D": "Overburden Bank Depth",
    "H_R_D": "Hardrock River Depth",
    "H_B_D": "Hardrock Bank Depth",
    "EE_1": "Executive Engineer_1",
    "EE_2": "Executive Engineer_2",
    "SDE_1": "Sub Divisional Engineer_1",
    "SDE_2": "Sub Divisional Engineer_2",
}
df_copy.rename(columns=COLUMN_RENAME_MAP, inplace=True)

# ---------------------------------------------------------------
# STEP 4: READ COMPUTED AMOUNTS
# ---------------------------------------------------------------
for i, file_path in enumerate(generated_files):
    wb = load_workbook(file_path, data_only=True)
    ws = wb[TARGET_SHEET]
    value = ws[TARGET_CELL].value
    df_copy.loc[i, "Amount"] = value
    print(f"Extracted Amount from {file_path}: {value}")

# ---------------------------------------------------------------
# STEP 5: WRITE FINAL data_copy.xlsx
# ---------------------------------------------------------------
temp_path = os.path.join(OUTPUT_FOLDER, "_temp_data_copy.xlsx")
df_copy.to_excel(temp_path, index=False)

wb = load_workbook(temp_path)
wb.save(data_copy_path)
wb.close()
os.remove(temp_path)

autofit_columns(data_copy_path, min_width=12)
print(f"\nUpdated data sheet saved at: {data_copy_path}")
