import pandas as pd
import geopandas as gpd
import os
import re
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

# ========================= CONFIGURATION =========================
EXCEL_FILE = r"C:\Users\Swapnali\Desktop\Test Upstream\Catchment Gross Yield (MCM).xlsx"
CATCHMENT_FOLDER = r"C:\Users\Swapnali\Desktop\Test Upstream\Ulhas & Vaitarna Catchments"
POINTS_FOLDER = r"C:\Users\Swapnali\Desktop\Test Upstream\Upstream Utilization\Ulhas_With_Upstream"

OUTPUT_EXCEL = "Ulhas_Gross_Yield_of_Catchments_with_upstream.xlsx"

# Desired column order for Summary sheet
SUMMARY_COLUMNS = [
    "Sr. No.",
    "Name of Lifting Points",
    "Gross Yield (MCM)",
    "Points Names",
    "Points Values",
    "Sum_Upstream_u",
    "Difference (Gross Yield - Sum Upstream_u)"
]
# =================================================================

def normalize_name(name):
    if pd.isna(name):
        return ""
    name = str(name).strip()
    normalized = re.sub(r'[^a-zA-Z0-9]', '', name.lower())
    return normalized

def autofit_and_center(ws):
    """Auto-fit columns and center-align all cells"""
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center_alignment
    
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 4, 60)
        ws.column_dimensions[column_letter].width = adjusted_width

# Step 1: Load points shapefile
points_shp_path = None
for file in os.listdir(POINTS_FOLDER):
    if file.lower().endswith(".shp"):
        points_shp_path = os.path.join(POINTS_FOLDER, file)
        break

if not points_shp_path:
    raise FileNotFoundError(f"No .shp file found in folder: {POINTS_FOLDER}")

print(f"Loading points shapefile: {points_shp_path}")
points_gdf = gpd.read_file(points_shp_path)

if "upstream_u" not in points_gdf.columns:
    raise KeyError(f"Column 'upstream_u' not found. Found: {list(points_gdf.columns)}")

points_gdf["upstream_u"] = pd.to_numeric(points_gdf["upstream_u"], errors='coerce').fillna(0)

# Step 2: Load Excel
print(f"Reading Excel: {EXCEL_FILE}")
df = pd.read_excel(EXCEL_FILE)
df["Gross Yield (MCM)"] = pd.to_numeric(df["Gross Yield (MCM)"], errors='coerce')

# Prepare expanded summary rows
summary_rows = []

print(f"\nScanning '{CATCHMENT_FOLDER}' for shapefiles...")
catchment_map = {}
for file in os.listdir(CATCHMENT_FOLDER):
    if file.lower().endswith(".shp"):
        stem = os.path.splitext(file)[0]
        norm_key = normalize_name(stem)
        full_path = os.path.join(CATCHMENT_FOLDER, file)
        catchment_map[norm_key] = full_path
        print(f"   ✓ Found: {file} → normalized: '{norm_key}'")

print(f"Total catchment shapefiles found: {len(catchment_map)}\n")

print(f"Processing {len(df)} catchments...\n")

for idx, row in df.iterrows():
    lifting_name_raw = str(row["Name of Lifting Points"])
    lifting_norm = normalize_name(lifting_name_raw)
    sr_no = row.get("Sr. No.", idx + 1)
    gross_yield = row["Gross Yield (MCM)"]
    
    shp_path = catchment_map.get(lifting_norm)
    
    if not shp_path:
        print(f"  ✗ NA - No matching shapefile for '{lifting_name_raw}'")
        summary_rows.append({
            "Sr. No.": sr_no,
            "Name of Lifting Points": lifting_name_raw,
            "Gross Yield (MCM)": gross_yield,
            "Points Names": "NA",
            "Points Values": None,
            "Sum_Upstream_u": "NA",
            "Difference (Gross Yield - Sum Upstream_u)": "NA"
        })
        continue
    
    try:
        catchment_gdf = gpd.read_file(shp_path)
    except Exception:
        print(f"  ✗ NA - Could not read shapefile for '{lifting_name_raw}'")
        summary_rows.append({
            "Sr. No.": sr_no,
            "Name of Lifting Points": lifting_name_raw,
            "Gross Yield (MCM)": gross_yield,
            "Points Names": "NA",
            "Points Values": None,
            "Sum_Upstream_u": "NA",
            "Difference (Gross Yield - Sum Upstream_u)": "NA"
        })
        continue
    
    # CRS handling
    if catchment_gdf.crs != points_gdf.crs and catchment_gdf.crs is not None and points_gdf.crs is not None:
        try:
            catchment_gdf = catchment_gdf.to_crs(points_gdf.crs)
        except:
            pass
    
    points_inside = gpd.sjoin(points_gdf, catchment_gdf, how="inner", predicate="intersects").copy()
    
    num_points = len(points_inside)
    total_upstream = float(points_inside["upstream_u"].sum()) if not points_inside.empty else 0.0
    diff = (gross_yield - total_upstream) if pd.notna(gross_yield) else "NA"
    
    print(f"  ✓ {lifting_name_raw}: {num_points} points → Sum = {total_upstream}")
    
    # Prepare point details - One row per point
    if not points_inside.empty:
        name_col = 'Name' if 'Name' in points_inside.columns else None
        point_names = [str(x) for x in points_inside[name_col].tolist()] if name_col else [f"Point_{i+1}" for i in range(num_points)]
        point_values = [float(x) for x in points_inside["upstream_u"].tolist()]
        
        for i in range(num_points):
            summary_rows.append({
                "Sr. No.": sr_no,
                "Name of Lifting Points": lifting_name_raw,
                "Gross Yield (MCM)": gross_yield,
                "Points Names": point_names[i],
                "Points Values": point_values[i],          # Numeric for calculations
                "Sum_Upstream_u": total_upstream,
                "Difference (Gross Yield - Sum Upstream_u)": diff
            })
    else:
        # Single row for catchments with zero points
        summary_rows.append({
            "Sr. No.": sr_no,
            "Name of Lifting Points": lifting_name_raw,
            "Gross Yield (MCM)": gross_yield,
            "Points Names": "",
            "Points Values": None,
            "Sum_Upstream_u": 0.0,
            "Difference (Gross Yield - Sum Upstream_u)": diff
        })

# Create Summary DataFrame with exact column order
summary_df = pd.DataFrame(summary_rows)
summary_df = summary_df[SUMMARY_COLUMNS]

# Step 3: Write to Excel (Only Summary sheet)
print("\nWriting Excel with merging and formatting...")
with pd.ExcelWriter(OUTPUT_EXCEL, engine='openpyxl') as writer:
    summary_df.to_excel(writer, sheet_name="Summary", index=False)

# Load workbook for merging cells
wb = load_workbook(OUTPUT_EXCEL)
ws = wb["Summary"]

# Find column indices
cols = {}
for c in range(1, ws.max_column + 1):
    header = str(ws.cell(1, c).value).strip()
    cols[header] = c

# Merge repeated columns per lifting point group
merge_cols = ["Sr. No.", "Name of Lifting Points", "Gross Yield (MCM)", "Sum_Upstream_u", "Difference (Gross Yield - Sum Upstream_u)"]

start_row = 2
while start_row <= ws.max_row:
    current_lifting = ws.cell(start_row, cols.get("Name of Lifting Points")).value
    end_row = start_row
    while end_row + 1 <= ws.max_row and ws.cell(end_row + 1, cols.get("Name of Lifting Points")).value == current_lifting:
        end_row += 1
    
    if end_row > start_row:
        for col_name in merge_cols:
            col_idx = cols.get(col_name)
            if col_idx:
                ws.merge_cells(start_row=start_row, start_column=col_idx, 
                              end_row=end_row, end_column=col_idx)
    
    start_row = end_row + 1

# Apply autofit and center alignment
autofit_and_center(ws)

wb.save(OUTPUT_EXCEL)

print("\n" + "="*90)
print("✅ Process completed successfully!")
print(f"Output file: {OUTPUT_EXCEL}")
print("\nSummary sheet now contains ALL data (no extra sheets):")
print("   • Exact column order as requested")
print("   • Points Values column is numeric (ready for calculations)")
print("   • Merged cells for repeated information per lifting point")
print("   • All columns auto-fitted and center-aligned")
print("\nYou can now perform numerical operations directly on the 'Points Values' column.")