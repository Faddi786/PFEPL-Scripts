import re
import os
import pandas as pd
import pdfplumber
from PIL import Image
import pytesseract
import io

# Configure tesseract path if needed (Windows example)
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'


def get_pdf_info():
    """Get PDF path and new date from user for each PDF"""
    pdf_path = input("\nEnter PDF file path (or 'done' to finish): ").strip().strip('"')
    
    if pdf_path.lower() == 'done':
        return None, None
    
    while not os.path.exists(pdf_path):
        print(f"File not found: {pdf_path}")
        pdf_path = input("Enter PDF file path (or 'done' to finish): ").strip().strip('"')
        if pdf_path.lower() == 'done':
            return None, None
    
    while True:
        new_date = input(f"Enter new date for {os.path.basename(pdf_path)} (format: DD-MM-YYYY): ").strip()
        if re.match(r'\d{2}-\d{2}-\d{4}', new_date):
            return pdf_path, new_date
        print("Invalid format. Please use DD-MM-YYYY (e.g., 15-04-2026)")


def extract_tracking_summary_dates_from_image(page):
    """Extract left and right dates from Tracking Summary table header using OCR"""
    try:
        # Convert page to image
        img = page.to_image(resolution=200).original
        text = pytesseract.image_to_string(img)
        
        # Look for pattern: date on left and date on right in the same line
        lines = text.split('\n')
        
        left_date = None
        right_date = None
        
        for line in lines:
            # Find all date-time patterns in the line
            dates = re.findall(r'(\d{2}-\d{2}-\d{4}\s+\d{2}:\d{2}:\d{2})', line)
            if len(dates) >= 2:
                left_date = dates[0]
                right_date = dates[1]
                break
            
            # Also try to find dates with just time (if date is elsewhere)
            if len(dates) == 1:
                # Look for time pattern
                times = re.findall(r'(\d{2}:\d{2}:\d{2})', line)
                if len(times) >= 2:
                    # Use the date from the first match
                    left_date = f"{dates[0].split()[0]} {times[0]}"
                    right_date = f"{dates[0].split()[0]} {times[1]}"
                    break
        
        return left_date, right_date
    except Exception as e:
        print(f"    OCR error for Tracking Summary: {e}")
        return None, None


def extract_residuals_dates_from_image(page):
    """Extract date-time from Residuals graph caption using OCR"""
    try:
        # Convert page to image
        img = page.to_image(resolution=200).original
        text = pytesseract.image_to_string(img)
        
        # Look for pattern like "31-07-2024 11:24:30 - 31-07-2024 15:58:30"
        lines = text.split('\n')
        
        for line in lines:
            # Find date-time range pattern
            match = re.search(r'(\d{2}-\d{2}-\d{4}\s+\d{2}:\d{2}:\d{2})\s*-\s*(\d{2}-\d{2}-\d{4}\s+\d{2}:\d{2}:\d{2})', line)
            if match:
                return match.group(1), match.group(2)
            
            # Also try with just times (if date appears separately)
            times = re.findall(r'(\d{2}:\d{2}:\d{2})', line)
            if len(times) >= 2:
                date_match = re.search(r'(\d{2}-\d{2}-\d{4})', line)
                if date_match:
                    date_str = date_match.group(1)
                    return f"{date_str} {times[0]}", f"{date_str} {times[1]}"
        
        return None, None
    except Exception as e:
        print(f"    OCR error for Residuals: {e}")
        return None, None


def find_section_pages(pdf_path):
    """Find all pages that have 'Tracking Summary' or 'Residuals' as the first text"""
    tracking_pages = []
    residuals_pages = []
    processing_style_info = []
    
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            
            # Get first non-empty line
            lines = [line.strip() for line in text.split('\n') if line.strip()]
            first_line = lines[0] if lines else ""
            
            # Check for Tracking Summary
            if first_line == "Tracking Summary" or (first_line.startswith("Tracking Summary") and len(first_line) < 30):
                tracking_pages.append(page_num)
                print(f"    Found Tracking Summary at page {page_num}")
            
            # Check for Residuals
            elif first_line == "Residuals" or (first_line.startswith("Residuals") and len(first_line) < 30):
                residuals_pages.append(page_num)
                print(f"    Found Residuals at page {page_num}")
            
            # Find Processing style pages (could be anywhere on page)
            if re.search(r'Processing\s+style', text, re.IGNORECASE):
                # Check if it's at the top (within first 10 lines)
                for line_idx, line in enumerate(lines[:10]):
                    if re.search(r'Processing\s+style', line, re.IGNORECASE):
                        is_at_top = True
                        processing_style_info.append({
                            'page_num': page_num,
                            'line_number': line_idx,
                            'is_at_top': is_at_top
                        })
                        print(f"    Found 'Processing style' at top of page {page_num}")
                        break
                else:
                    # Not at top (in middle/bottom)
                    processing_style_info.append({
                        'page_num': page_num,
                        'line_number': 99,
                        'is_at_top': False
                    })
                    print(f"    Found 'Processing style' in middle/bottom of page {page_num}")
    
    return tracking_pages, residuals_pages, processing_style_info


def get_datetime_for_page(pdf_path, page_num, section_type):
    """Extract datetime from a specific page using OCR on the image"""
    with pdfplumber.open(pdf_path) as pdf:
        page = pdf.pages[page_num - 1]
        
        if section_type == 'tracking':
            return extract_tracking_summary_dates_from_image(page)
        else:  # residuals
            return extract_residuals_dates_from_image(page)


def get_user_selected_pages(residuals_sections, total_pages):
    """Ask user to select additional pages to include as Residuals sections"""
    # Calculate end page numbers + 1 for each Residuals section
    end_plus_one_pages = []
    for r in residuals_sections:
        end_plus_one = r['end_page'] + 1
        if end_plus_one <= total_pages:
            end_plus_one_pages.append(end_plus_one)
    
    if not end_plus_one_pages:
        print("\n    No pages available after Residuals sections.")
        return []
    
    print(f"\n    Pages after Residuals sections (end_page + 1): {end_plus_one_pages}")
    print("    These pages may contain additional graphs.")
    
    user_input = input("\n    Enter page numbers to include as Residuals (comma-separated, or press Enter to skip): ").strip()
    
    if not user_input:
        return []
    
    # Parse user input
    selected_pages = []
    for part in user_input.split(','):
        part = part.strip()
        if part.isdigit():
            page = int(part)
            if 1 <= page <= total_pages:
                selected_pages.append(page)
            else:
                print(f"    Warning: Page {page} is out of range (1-{total_pages})")
        else:
            print(f"    Warning: '{part}' is not a valid page number")
    
    return sorted(set(selected_pages))


def process_pdf(pdf_path, new_date):
    """Process a single PDF using the page detection logic"""
    pdf_name = os.path.basename(pdf_path)
    base_name = os.path.splitext(pdf_name)[0]
    
    print(f"\n  Scanning PDF: {pdf_name}")
    
    # Find all section start pages and processing style info
    tracking_pages, residuals_pages, processing_style_info = find_section_pages(pdf_path)
    
    # Get total pages
    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)
    
    tracking_sections = []
    residuals_sections = []
    
    # Process Tracking Summary sections
    for i, start_page in enumerate(tracking_pages):
        # Find end page: next Residuals page - 1, or total_pages if no next Residuals
        next_residuals = None
        for rp in residuals_pages:
            if rp > start_page:
                next_residuals = rp
                break
        
        end_page = (next_residuals - 1) if next_residuals else total_pages
        
        # Get datetime from this page using OCR
        print(f"    Extracting dates from Tracking Summary page {start_page}...")
        start_dt, end_dt = get_datetime_for_page(pdf_path, start_page, 'tracking')
        
        tracking_sections.append({
            'start_page': start_page,
            'end_page': end_page,
            'start_datetime': start_dt,
            'end_datetime': end_dt
        })
        
        print(f"    Tracking Summary: pages {start_page}-{end_page} | {start_dt} -> {end_dt}")
    
    # Process Residuals sections
    for i, start_page in enumerate(residuals_pages):
        # Find end page: next Processing style page based on position
        next_processing = None
        next_processing_at_top = False
        
        for p_info in processing_style_info:
            if p_info['page_num'] > start_page:
                next_processing = p_info['page_num']
                next_processing_at_top = p_info['is_at_top']
                break
        
        # Determine end page based on Processing style position
        if next_processing:
            if next_processing_at_top:
                # Processing style is at top of page - subtract 1
                end_page = next_processing - 1
                print(f"    Processing style at top of page {next_processing} - using page {end_page} as end")
            else:
                # Processing style is in middle - use current page as end
                end_page = next_processing
                print(f"    Processing style in middle of page {next_processing} - using same page as end")
        else:
            end_page = total_pages
        
        # Get datetime from this page using OCR
        print(f"    Extracting dates from Residuals page {start_page}...")
        start_dt, end_dt = get_datetime_for_page(pdf_path, start_page, 'residuals')
        
        residuals_sections.append({
            'start_page': start_page,
            'end_page': end_page,
            'start_datetime': start_dt,
            'end_datetime': end_dt
        })
        
        print(f"    Residuals: pages {start_page}-{end_page} | {start_dt} -> {end_dt}")
    
    # Ask user for additional pages to include as Residuals
    print(f"\n  Checking for additional graph pages...")
    additional_pages = get_user_selected_pages(residuals_sections, total_pages)
    
    for page_num in additional_pages:
        print(f"\n    Processing additional page {page_num} as Residuals...")
        
        # For additional pages, start and end page are the same (single page section)
        start_dt, end_dt = get_datetime_for_page(pdf_path, page_num, 'residuals')
        
        residuals_sections.append({
            'start_page': page_num,
            'end_page': page_num,
            'start_datetime': start_dt,
            'end_datetime': end_dt
        })
        
        print(f"    Additional Residuals: page {page_num} | {start_dt} -> {end_dt}")
    
    # Sort residuals sections by start page
    residuals_sections.sort(key=lambda x: x['start_page'])
    
    return tracking_sections, residuals_sections, base_name, pdf_name, additional_pages


def create_excel_for_pdf(pdf_path, new_date):
    """Process a single PDF and create its Excel file"""
    tracking_sections, residuals_sections, base_name, pdf_name, additional_pages = process_pdf(pdf_path, new_date)
    
    # Prepare Tracking Summary DataFrame
    tracking_data = []
    for t in tracking_sections:
        start_modified = None
        end_modified = None
        
        if t['start_datetime']:
            # Extract time only from full datetime
            time_match = re.search(r'(\d{2}:\d{2}:\d{2})', str(t['start_datetime']))
            if time_match:
                start_modified = f"{new_date} {time_match.group(1)}"
        
        if t['end_datetime']:
            time_match = re.search(r'(\d{2}:\d{2}:\d{2})', str(t['end_datetime']))
            if time_match:
                end_modified = f"{new_date} {time_match.group(1)}"
        
        tracking_data.append({
            'PDF File Name': pdf_name,
            'New Date Applied': new_date,
            'Start Page': t['start_page'],
            'End Page': t['end_page'],
            'Original Start Date-Time': t['start_datetime'],
            'Original End Date-Time': t['end_datetime'],
            'Modified Start Date-Time': start_modified,
            'Modified End Date-Time': end_modified
        })
    
    # Prepare Residuals DataFrame
    residuals_data = []
    for r in residuals_sections:
        start_modified = None
        end_modified = None
        merged_range = None
        
        if r['start_datetime']:
            time_match = re.search(r'(\d{2}:\d{2}:\d{2})', str(r['start_datetime']))
            if time_match:
                start_modified = f"{new_date} {time_match.group(1)}"
        
        if r['end_datetime']:
            time_match = re.search(r'(\d{2}:\d{2}:\d{2})', str(r['end_datetime']))
            if time_match:
                end_modified = f"{new_date} {time_match.group(1)}"
        
        if start_modified and end_modified:
            merged_range = f"{start_modified} - {end_modified}"
        
        residuals_data.append({
            'PDF File Name': pdf_name,
            'New Date Applied': new_date,
            'Start Page': r['start_page'],
            'End Page': r['end_page'],
            'Original Start Date-Time': r['start_datetime'],
            'Original End Date-Time': r['end_datetime'],
            'Modified Start Date-Time': start_modified,
            'Modified End Date-Time': end_modified,
            'Merged Date-Time Range': merged_range
        })
    
    # Create DataFrames
    df_tracking = pd.DataFrame(tracking_data)
    df_residuals = pd.DataFrame(residuals_data)
    
    # Write to Excel in script directory
    script_dir = os.path.dirname(os.path.abspath(__file__)) or os.getcwd()
    output_path = os.path.join(script_dir, f"{base_name}.xlsx")
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_tracking.to_excel(writer, sheet_name='Tracking Summary', index=False)
        if not df_residuals.empty:
            df_residuals.to_excel(writer, sheet_name='Residuals', index=False)
    
    # Auto-adjust column widths
    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    wb.save(output_path)
    
    # Calculate counts for display
    auto_count = len([r for r in residuals_sections if r['start_page'] not in additional_pages])
    user_count = len(additional_pages)
    
    print(f"\n  ✓ Excel created: {output_path}")
    print(f"    - Tracking Summary sections: {len(tracking_data)}")
    print(f"    - Residuals sections: {len(residuals_data)}")
    print(f"      - Auto-detected Residuals: {auto_count}")
    print(f"      - User-added Residuals: {user_count}")
    
    return output_path


def main():
    """Main function to process PDFs one by one"""
    print("=" * 70)
    print("PDF Baseline Processing Report Extractor with OCR")
    print("=" * 70)
    print("\nThis script uses OCR to extract date-time information from:")
    print("  - Tracking Summary: Dates from table headers (left and right)")
    print("  - Residuals: Date-time ranges from graph captions")
    print("\nAfter processing, you can manually add additional graph pages.")
    print("Enter 'done' when you have processed all PDFs.\n")
    
    processed_files = []
    
    while True:
        pdf_path, new_date = get_pdf_info()
        
        if pdf_path is None:
            break
        
        print(f"\n{'='*50}")
        print(f"Processing: {os.path.basename(pdf_path)}")
        print(f"New date: {new_date}")
        print(f"{'='*50}")
        
        try:
            output_path = create_excel_for_pdf(pdf_path, new_date)
            processed_files.append({
                'pdf': os.path.basename(pdf_path),
                'excel': os.path.basename(output_path),
                'date': new_date
            })
        except Exception as e:
            print(f"\n  ✗ Error processing {pdf_path}: {e}")
            import traceback
            traceback.print_exc()
            print("    Skipping this file...")
    
    # Summary
    print("\n" + "=" * 70)
    print("PROCESSING SUMMARY")
    print("=" * 70)
    
    if processed_files:
        print(f"\nSuccessfully processed {len(processed_files)} PDF(s):\n")
        for i, file_info in enumerate(processed_files, 1):
            print(f"{i}. PDF: {file_info['pdf']}")
            print(f"   Excel: {file_info['excel']}")
            print(f"   Date applied: {file_info['date']}\n")
    else:
        print("\nNo files were processed.")
    
    print("=" * 70)


if __name__ == "__main__":
    main()