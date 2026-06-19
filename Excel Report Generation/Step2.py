# Step2.py
from openpyxl import load_workbook, Workbook
from openpyxl.cell.cell import MergedCell
from copy import copy
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def set_page_setup(ws: Worksheet):
    """Apply A3, Landscape, Narrow margins"""
    # A3 paper
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

    ws.page_margins.left = 0.2952756
    ws.page_margins.right = 0.2952756
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3


def copy_column_widths(src_ws, dst_ws, start_col=2):
    """Copies column widths from src_ws (starting column B) to dst_ws (starting column A)"""
    for col_idx in range(start_col, src_ws.max_column + 1):
        src_letter = get_column_letter(col_idx)
        dst_letter = get_column_letter(col_idx - 1)

        if src_letter in src_ws.column_dimensions:
            dst_ws.column_dimensions[dst_letter].width = \
                src_ws.column_dimensions[src_letter].width


def copy_table(src_ws, dst_ws, start, end, repeat_rows):
    """Copies table data, formatting, merged cells, column widths, and repeat rows"""

    # ---- Copy values + formatting ----
    for r in range(start, end + 1):
        for c in range(2, src_ws.max_column + 1):  # skip column A
            src = src_ws.cell(r, c)
            dst = dst_ws.cell(row=r - start + 1, column=c - 1)

            if isinstance(dst, MergedCell):
                continue

            dst.value = src.value

            if src.has_style:
                dst.font = copy(src.font)
                dst.fill = copy(src.fill)
                dst.border = copy(src.border)
                dst.alignment = copy(src.alignment)

    # ---- Copy merged cells ----
    for m in src_ws.merged_cells.ranges:
        if m.min_row >= start and m.max_row <= end:
            dst_ws.merge_cells(
                start_row=m.min_row - start + 1,
                end_row=m.max_row - start + 1,
                start_column=m.min_col - 1,
                end_column=m.max_col - 1
            )

    # ---- Column widths ----
    copy_column_widths(src_ws, dst_ws, start_col=2)

    # ---- Repeat rows (Print Titles) ----
    if repeat_rows:
        rel_rows = [r - start + 1 for r in repeat_rows]
        dst_ws.print_title_rows = f"{min(rel_rows)}:{max(rel_rows)}"


def main():
    input_wb = load_workbook("input.xlsx", data_only=True)
    out_wb = Workbook()

    # Remove default empty sheet
    out_wb.remove(out_wb.active)

    table_idx = 1

    for ws in input_wb.worksheets:
        start = None
        repeat_rows = []

        for r in range(1, ws.max_row + 1):
            raw = ws.cell(r, 1).value
            text = str(raw).lower() if raw else ""

            if "start" in text:
                start = r

            if "repeat" in text:
                repeat_rows.append(r)

            if "end" in text and start:
                new_ws = out_wb.create_sheet(title=f"Table_{table_idx}")

                # Apply page setup FIRST
                set_page_setup(new_ws)

                copy_table(ws, new_ws, start, r, repeat_rows)

                table_idx += 1
                start = None
                repeat_rows = []

    out_wb.save("output.xlsx")
    print("✔ Step 2 completed: tables copied, A3 landscape, narrow margins applied")


if __name__ == "__main__":
    main()
