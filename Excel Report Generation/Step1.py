# Step1.py
from openpyxl import load_workbook

KEYWORDS = ("start", "end")

def normalize(cell):
    if not cell:
        return []
    return [x.strip().lower() for x in str(cell).split("/")]

def validate_sheet(ws):
    in_table = False

    for r in range(1, ws.max_row + 1):
        words = normalize(ws.cell(r, 1).value)

        if "start" in words:
            if in_table:
                raise ValueError(f"{ws.title}: START without END at row {r}")
            in_table = True

        if "end" in words:
            if not in_table:
                raise ValueError(f"{ws.title}: END without START at row {r}")
            in_table = False

    if in_table:
        raise ValueError(f"{ws.title}: START without matching END")

def main():
    wb = load_workbook("input.xlsx", data_only=True)

    print("\nVALIDATION\n----------")
    for ws in wb.worksheets:
        validate_sheet(ws)
        print(f"✔ {ws.title}: OK")

    print("\n✔ Validation completed successfully")

if __name__ == "__main__":
    main()
