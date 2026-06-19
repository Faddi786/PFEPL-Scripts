# Step3.py (CONFIG-DRIVEN, NON-INTERACTIVE)

import json
import sys
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

OUTPUT_FILE = "output.xlsx"
CONFIG_FILE = "section_config.json"


def setup_page(ws):
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.horizontalCentered = True
    ws.page_setup.verticalCentered = True
    ws.sheet_view.view = "pageLayout"


def create_section_sheet(wb, title, text):
    ws = wb.create_sheet(title=title)

    cell = ws.cell(row=1, column=1)
    cell.value = text

    cell.font = Font(
        name="Cambria",
        size=30,
        bold=False
    )

    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=False
    )

    autofit_column_for_large_font(
    ws=ws,
    col_idx=1,
    text=text,
    font_size=cell.font.size)

    # Excel row height ≈ font size × 1.4
    ws.row_dimensions[1].height = cell.font.size * 1.4

    setup_page(ws)
    return ws

    
    # # -----------------------------
    # # AUTO-FIT COLUMN (Excel-like)
    # # -----------------------------
    # text_length = len(str(text))
    
    # # Excel approx: 1 char ≈ 1 unit width
    # # Cap width to avoid insane columns
    # adjusted_width = min(text_length + 4, 1200)
    # ws.column_dimensions[get_column_letter(1)].width = adjusted_width

    # -----------------------------
    # AUTO-FIT ROW (Excel-like)
    # -----------------------------


def autofit_column_for_large_font(ws, col_idx, text, font_size):
    """
    Excel column width units assume ~Calibri 11.
    Scale width proportionally for large fonts.
    """

    base_char_width = 1.0          # Excel unit per char at size 11
    font_scale = font_size / 11.0  # scale for font size
    padding = 4

    width = (len(text) * base_char_width * font_scale) + padding

    # Excel hard limit
    width = min(width, 255)

    ws.column_dimensions[get_column_letter(col_idx)].width = width


def load_config():
    with open(CONFIG_FILE, "r", encoding="utf-8") as f:
        cfg = json.load(f)

    if "sections" not in cfg or not isinstance(cfg["sections"], list):
        raise ValueError("Invalid config: 'sections' list missing")

    return cfg["sections"]


def validate_section(entry, wb):
    if entry.get("position") not in ("before", "after"):
        raise ValueError(f"Invalid position: {entry}")

    table = entry.get("table")
    if not table or table not in wb.sheetnames:
        raise ValueError(f"Target table not found: {table}")

    text = entry.get("text")
    if not text or not text.strip():
        raise ValueError(f"Empty section text for table {table}")


def remove_existing_sections(wb):
    for ws in list(wb.worksheets):
        if ws.title.lower().startswith("section_"):
            wb.remove(ws)


def main():
    wb = load_workbook(OUTPUT_FILE)

    sections = load_config()

    remove_existing_sections(wb)

    section_count = 1

    for entry in sections:
        validate_section(entry, wb)

        section_ws = create_section_sheet(
            wb,
            title=f"Section_{section_count}",
            text=entry["text"]
        )

        target = entry["table"]
        target_index = wb.sheetnames.index(target)

        # Remove and reinsert to control position
        wb._sheets.remove(section_ws)

        if entry["position"] == "before":
            wb._sheets.insert(target_index, section_ws)
        else:
            wb._sheets.insert(target_index + 1, section_ws)

        section_count += 1

    # Enforce layout everywhere
    for ws in wb.worksheets:
        setup_page(ws)

    wb.save(OUTPUT_FILE)
    print("✔ Step 3 completed successfully (config-driven)")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print("ERROR in Step3:", e)
        sys.exit(1)
