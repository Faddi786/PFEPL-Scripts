import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# --- Configuration ---     
PAGE_WIDTH_TARGET = 155   # A3 landscape target width
MIN_COL_WIDTH = 5
SHRINK_FACTOR = 0.97
MAX_ITERATIONS = 20 

import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

INPUT_FILE = os.path.join(BASE_DIR, "output.xlsx")
OUTPUT_FILE = os.path.join(BASE_DIR, "output_formatted.xlsx")

if not os.path.exists(INPUT_FILE):
    raise FileNotFoundError(f"Missing input file: {INPUT_FILE}")

def get_max_content_width(sheet, col_idx):
    """Calculates max string length in a column, ignoring merged cells."""
    max_len = 0
    merged_ranges = sheet.merged_cells.ranges
    
    for row in range(1, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=col_idx)
        # Handle merged cells gracefully: ignore them for individual column width stats
        is_merged = any(cell.coordinate in rng for rng in merged_ranges)
        
        if not is_merged and cell.value:
            val_len = len(str(cell.value))
            if val_len > max_len:
                max_len = val_len
    return max_len if max_len > 0 else 8.43

def apply_narrow_margins_to_all_sheets(wb):
    for ws in wb.worksheets:
        ws.page_margins.top = 0.984252
        ws.page_margins.bottom = 0.984252
        ws.page_margins.left = 0.2952756
        ws.page_margins.right = 0.2952756
        ws.page_margins.header = 0.19685
        ws.page_margins.footer = 0.19685

def set_a3_only(sheet):
    """Only enforce A3 paper size, nothing else"""
    sheet.page_setup.paperSize = 8  # A3

def get_total_width(sheet):
    total = 0
    for col_idx in range(1, sheet.max_column + 1):
        col_letter = get_column_letter(col_idx)
        w = sheet.column_dimensions[col_letter].width
        total += w if w else 8.43
    return total

def apply_reduction(sheet, col_indices, steps):
    """Helper to reduce specific columns by the shrink factor N times."""
    for col_idx in col_indices:
        col_letter = get_column_letter(col_idx)
        dim = sheet.column_dimensions[col_letter]
        current_w = dim.width or 8.43
        new_w = current_w * (SHRINK_FACTOR ** steps)
        dim.width = max(new_w, MIN_COL_WIDTH)

def tiered_shrink_logic(sheet):
    iteration = 0
    while iteration < MAX_ITERATIONS:
        if get_total_width(sheet) <= PAGE_WIDTH_TARGET:
            return

        # 1. Map columns to their max content length
        col_stats = []
        for c in range(1, sheet.max_column + 1):
            col_stats.append({
                'idx': c,
                'content_len': get_max_content_width(sheet, c)
            })

        # Sort columns by content length (descending)
        col_stats.sort(key=lambda x: x['content_len'], reverse=True)
        total_cols = len(col_stats)
        
        # Define Tiers
        tier1 = [c['idx'] for c in col_stats[:int(total_cols * 0.4)]]
        tier2 = [c['idx'] for c in col_stats[int(total_cols * 0.4):int(total_cols * 0.7)]]
        tier3 = [c['idx'] for c in col_stats[int(total_cols * 0.7):]]

        # Tier 1 Reduction (Top 40% - ~10% reduction)
        apply_reduction(sheet, tier1, 3)
        if get_total_width(sheet) <= PAGE_WIDTH_TARGET: return

        # Tier 2 Reduction (Next 30% - ~6% reduction)
        apply_reduction(sheet, tier2, 2)
        if get_total_width(sheet) <= PAGE_WIDTH_TARGET: return

        # Tier 3 Reduction (Next 30% - ~3% reduction)
        apply_reduction(sheet, tier3, 1)
        if get_total_width(sheet) <= PAGE_WIDTH_TARGET: return

        iteration += 1

def format_excel_a3_tiered(file_path, output_path):
    wb = openpyxl.load_workbook(file_path)
    
    if "Optimization_Log" not in wb.sheetnames:
        print("Log sheet missing.")
        return

    log_ws = wb["Optimization_Log"]
    
    # Identify qualifying sheets: Table_* name AND Final Pages Wide > 1
    sheets_to_fix = []
    for row in log_ws.iter_rows(min_row=2, values_only=True):
        name, _, pages_wide = row[0], row[1], row[2]
        if name and str(name).startswith("Table_") and pages_wide and pages_wide > 1:
            sheets_to_fix.append(name)

    for sheet_name in sheets_to_fix:
        if sheet_name not in wb.sheetnames: continue
        sheet = wb[sheet_name]

        # Setup A3 Landscape
        sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
        sheet.page_setup.paperSize = 8 
        sheet.page_setup.fitToWidth = None
        sheet.page_setup.fitToHeight = None
        sheet.sheet_properties.pageSetUpPr.fitToPage = False

        # Run Tiered Shrinking
        tiered_shrink_logic(sheet)

        # Apply wrap text to all cells
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Auto-fit row heights
        for r in range(1, sheet.max_row + 1):
            sheet.row_dimensions[r].height = None
    
    if "Optimization_Log" in wb.sheetnames:
        del wb["Optimization_Log"]

    # ---- Apply A3 ONLY to non-Table sheets ----
    for ws in wb.worksheets:
        if not ws.title.startswith("Table_"):
            ws.page_setup.paperSize = 8  # A3 only

    # 🔒 Enforce Excel "Narrow" margins on ALL sheets
    apply_narrow_margins_to_all_sheets(wb)

    wb.save(output_path)
    print("A3 Tiered Formatting Complete.")

# if __name__ == "__main__":
#     format_excel_a3_tiered("output.xlsx", "output_formatted.xlsx")

if __name__ == "__main__":
    format_excel_a3_tiered(INPUT_FILE, OUTPUT_FILE)
