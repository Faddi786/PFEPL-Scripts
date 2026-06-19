import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ───────────────────────────────────────────────
# CONFIG
# ───────────────────────────────────────────────
input_file  = 'Theissen Polygons_Calculate percent & total area.xlsx'
output_file = 'Subbasin_Report_with_Percentage.xlsx'

# ───────────────────────────────────────────────
# 1. Read and prepare data
# ───────────────────────────────────────────────
df = pd.read_excel(input_file)  # Read with headers from row 0
# Select columns and drop only the header row (row 0), keep row 1 as it contains data
df = df[['name_1', 'Name', 'Shape_Area']].iloc[1:].reset_index(drop=True)
df.columns = ['Subbasin', 'Station', 'Area_sqkm']
df = df.dropna(how='all')  # Remove any completely empty rows
# Convert Area_sqkm to numeric, coercing errors to NaN
df['Area_sqkm'] = pd.to_numeric(df['Area_sqkm'], errors='coerce')
df = df.dropna(subset=['Area_sqkm'])  # Remove rows where area couldn't be converted

# Calculate total area per subbasin
subbasin_totals = df.groupby('Subbasin')['Area_sqkm'].sum().reset_index()
subbasin_totals = subbasin_totals.rename(columns={'Area_sqkm': 'Total_sqkm'})

# Merge total back to original rows
df = df.merge(subbasin_totals, on='Subbasin', how='left')

# Don't calculate percentage - leave it empty
df['Percentage'] = None

# Extract numeric part from Subbasin for proper numeric sorting (e.g., "Subbasin-1" -> 1, "Subbasin-10" -> 10)
def extract_subbasin_number(subbasin):
    match = re.search(r'(\d+)$', str(subbasin))
    return int(match.group(1)) if match else 0

df['Subbasin_Num'] = df['Subbasin'].apply(extract_subbasin_number)

# Sort subbasins numerically (1, 2, 3, 10, 11...) and within each subbasin by area descending
df = df.sort_values(['Subbasin_Num', 'Area_sqkm'], ascending=[True, False])
df = df.drop('Subbasin_Num', axis=1)  # Remove helper column

# ───────────────────────────────────────────────
# 2. Create output Excel
# ───────────────────────────────────────────────
wb = Workbook()
ws_report = wb.active
ws_report.title = "Report"

# ──────── Pretty Report sheet ────────
headers = ['Sr.No.', 'Sub Basin Name', 'Station Name', 'Station area (sq.km)', 'Total area (sq.km)', 'Percentage']
for col_idx, header in enumerate(headers, 1):
    cell = ws_report.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

current_row = 2
sr_no = 1
prev_subbasin = None
subbasin_start_row = None  # Track start row for each subbasin group

# First pass: write all data
for _, row in df.iterrows():
    sub = row['Subbasin']
    station = row['Station']
    area = row['Area_sqkm']
    total = row['Total_sqkm']

    # Check if this is the first row of a new subbasin
    is_first_row = (sub != prev_subbasin)

    if is_first_row:
        # If we had a previous subbasin, merge its cells before starting new one
        if prev_subbasin is not None and subbasin_start_row is not None:
            end_row = current_row - 1
            # Merge Subbasin column (B) only - NOT Sr.No.
            if end_row > subbasin_start_row:
                ws_report.merge_cells(f'B{subbasin_start_row}:B{end_row}')
                ws_report[f'B{subbasin_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
            # Merge Total area column (E)
            if end_row > subbasin_start_row:
                ws_report.merge_cells(f'E{subbasin_start_row}:E{end_row}')
                ws_report[f'E{subbasin_start_row}'].alignment = Alignment(horizontal='right', vertical='center')
        
        # New subbasin block - write Subbasin name
        ws_report.cell(row=current_row, column=2, value=sub)
        ws_report.cell(row=current_row, column=2).alignment = Alignment(horizontal='center', vertical='center')
        subbasin_start_row = current_row  # Track start of this subbasin group
        prev_subbasin = sub
    
    # Write Sr.No. for EVERY row (sequential: 1, 2, 3, 4...)
    ws_report.cell(row=current_row, column=1, value=sr_no)
    ws_report.cell(row=current_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
    sr_no += 1
    
    # If not first row of subbasin, leave Subbasin empty (will be merged later)
    if not is_first_row:
        ws_report.cell(row=current_row, column=2, value=None)

    # Station and values
    ws_report.cell(row=current_row, column=3, value=station)
    ws_report.cell(row=current_row, column=4, value=area)
    # Show total only on first row of each subbasin, leave percentage empty
    if is_first_row:
        ws_report.cell(row=current_row, column=5, value=total)
        ws_report.cell(row=current_row, column=5).alignment = Alignment(horizontal='right', vertical='center')
    else:
        ws_report.cell(row=current_row, column=5, value=None)
    ws_report.cell(row=current_row, column=6, value=None)  # Leave percentage empty

    current_row += 1

# Merge cells for the last subbasin group
if subbasin_start_row is not None:
    end_row = current_row - 1
    if end_row > subbasin_start_row:
        # Merge Subbasin column (B) only - NOT Sr.No.
        ws_report.merge_cells(f'B{subbasin_start_row}:B{end_row}')
        ws_report[f'B{subbasin_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        # Merge Total area column (E)
        ws_report.merge_cells(f'E{subbasin_start_row}:E{end_row}')
        ws_report[f'E{subbasin_start_row}'].alignment = Alignment(horizontal='right', vertical='center')

# Formatting - apply borders and number formats
from openpyxl.cell.cell import MergedCell
for row in ws_report.iter_rows(min_row=1, max_row=current_row-1, min_col=1, max_col=6):
    for cell in row:
        # Skip MergedCell objects (they are part of merged ranges)
        if not isinstance(cell, MergedCell):
            cell.border = thin_border
            if hasattr(cell, 'column_letter'):
                if cell.column_letter in ['D','E']:
                    cell.number_format = '0.000000'
                if cell.column_letter == 'F':
                    cell.number_format = '0.00%'

# Alignments & widths
ws_report.column_dimensions['A'].width = 8
ws_report.column_dimensions['B'].width = 18
ws_report.column_dimensions['C'].width = 22
ws_report.column_dimensions['D'].width = 16
ws_report.column_dimensions['E'].width = 16
ws_report.column_dimensions['F'].width = 12

for cell in ws_report['A']:
    cell.alignment = Alignment(horizontal='center')
for row in ws_report['D:F']:
    for cell in row:
        cell.alignment = Alignment(horizontal='right')

# Freeze header
ws_report.freeze_panes = 'A2'

# ───────────────────────────────────────────────
# Save
# ───────────────────────────────────────────────
wb.save(output_file)
print(f"Output saved to: {output_file}")
print(f"• Sheet 'Report' -> formatted view (like your example)")